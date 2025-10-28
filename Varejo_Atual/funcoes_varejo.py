from copy import deepcopy
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

def get_pivot_table(tb_analitica):
    return (
        pd.pivot_table(tb_analitica, index=['HISTORICO_BP_MIS', 'SUBTIPO', 'SEGMENTO', 'LINHA'], columns=['MES'], values=['VALOR'], aggfunc='sum', fill_value=0)
        .droplevel(level=0, axis=1).rename_axis(columns=None).reset_index()
    )

def gerar_tabela_raw(sheet_name):
    if sheet_name == 'R_MODELO':
        n_rows = 10019
        use_cols = 'A:AV'
    elif sheet_name == 'P_MODELO':
        n_rows = 3669
        use_cols = 'A:AW'
    elif sheet_name == 'S_MODELO':
        n_rows = 969
        use_cols = 'A:AW'
    elif sheet_name == 'Casc_MODELO':
        n_rows = 128
        use_cols = 'F:G'
    df = pd.read_excel('Modelo_BP_Atual.xlsx', engine='openpyxl', sheet_name=sheet_name, skiprows=9, nrows=n_rows, usecols=use_cols)
    return df

def tratar_cols_spread(row, r_col, p_col, param_col, dic_dias_mes, dic_du_mes, anomes, ano):
    '''Calcula o spread varrendo linha a linha do DataFrame, passar como lambda function.
    
    Parameters:
    row(linha): Linha do DataFrame
    r_col(str): String com o nome da Coluna oriunda do RModelo para cálculo do spread
    p_col(str): String com o nome da Coluna oriunda do PModelo para cálculo do spread
    param_col(str): String com o nome da Coluna com flag se o cálculo será feito com dias corridos ou úteis oriunda do SModelo, para cálculo do spread
    dic_dias_mes(dict): Meses e Anos e seus números de dias corridos, oriundo do arquivo de variáveis (variaveis_var_atual)
    dic_du_mes(dict): Meses e Anos e seus números de dias úteis, oriundo do arquivo de variáveis (variaveis_var_atual)
    anomes(str): coluna a ser criada com o resultado do spread para aquele anomes
    ano(str): ano referente ao cálculo do spread
    
    Returns:
    row[anomes](list): coluna com linhas preenchidas
    '''    
    if row[param_col] == 1 and row[p_col] > 0:
        row[anomes] = (((row[r_col]) / (dic_dias_mes.get(anomes)) * (dic_dias_mes.get(ano)))) / row[p_col]
    elif (row[param_col] == 0 or pd.isna(row[param_col]) == True) and row[p_col] > 0:
        row[anomes] = (((row[r_col]) / (dic_du_mes.get(anomes)) * (dic_du_mes.get(ano)))) / row[p_col]    
    else:
        row[anomes] = 0 
    return row[anomes]

def tratar_col_casc(row, col, col_cod):
    '''Ajusta algumas linhas 'Itens / Período', com base no Cod que serão usados nos joins entre tabelas Cascada e RModelo varrendo linha a linha do DataFrame, passar como lambda function.
    
    Parameters:
    row(linha): Linha do DataFrame
    col(str): String com o nome da Coluna a ser ajustada
    col_cod(str): String com o nome da Coluna Cod
    
    Returns:
    row[anomes](list): coluna com linhas preenchidas
    '''    
    # as linhas com os códigos a seguir terão sua coluna Itens / Período modificada
    if row[col_cod] in (13, 14, 15, 16, 17, 18, 19, 20, 22, 24, 25, 26, 27, 28):
        row[col] = row[col] + ' Margem'      
    elif row[col_cod] in (41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 59, 60, 61, 62, 63, 64, 65):
        row[col] = row[col] + ' Comissão'
    else:
        row[col] = row[col]
        
    return row[col] 

def ajustar_row_cod_0(modelo):
    '''Substitui as linhas que tem Cod 0 por espaço em branco.
    Parameters:
    modelo(Pandas DataFrame): Modelo a ser ajustado
    Returns:
    modelo(Pandas DataFrame): Modelo ajustado
    '''
    if isinstance(modelo, list):
        for df in modelo:
            df['Cod'] = df['Cod'].replace({'0': '', 0: ''})
    else:
        modelo['Cod'] = modelo['Cod'].replace({'0': '', 0: ''})
    return modelo

def gerar_tabela_rmodelo(rmodelo_raw, pivot_resultado, subtipos, segmentos, num_anos, cenario):
    '''Constrói o RModelo do Varejo Atual com Total, Contábil e Fictício.
    
    Parameters:
    rmodelo_raw(Pandas DataFrame): Template do RModelo com a estrutura (arquivo em Excel Modelo BP Atual)
    pivot_resultado(Pandas DataFrame): Tabela dinâmica oriunda da função gerar_pivot_table()
    subtipos(list): Subtipos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    segmentos(dict): Segmentos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    
    Returns:
    df_list_final(pandas DataFrame): RModelo final preenchido
    '''
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    print(y0, y1, y01, y02, y02, y03)

    # Pegar apenas as linhas e colunas necessárias para utilizar no preenchimento do RModelo
    rmodelo_raw = rmodelo_raw.iloc[1:]
    rmodelo_raw = rmodelo_raw.drop(rmodelo_raw.iloc[:, 12:40], axis = 1)
    rmodelo_raw = rmodelo_raw.rename(columns={'Unnamed: 9': '', '%': f'Reparto % {y1}', '%.1': f'Reparto % {y0}'})

    rmodelo_raw = rmodelo_raw[['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos', 'Linha excel', '', f'Reparto % {y1}',f'Reparto % {y0}', 'ISS', 'PIS', 'IR']]

    rmodelo_raw.dropna(subset=['Cod'])
    rmodelo_raw['ID'] = rmodelo_raw.index

    # Transformar o tipo das colunas Conceito e Cod para string do RModelo, além de substituir algumas strings para os joins
    rmodelo_raw['Conceito'] = rmodelo_raw['Conceito'].astype(str)
    rmodelo_raw['Cod'] = rmodelo_raw['Cod'].astype(str)
#     rmodelo_raw['Conceito'] = rmodelo_raw['Conceito'].replace('Contábil', 'Resultado contabil')
#     rmodelo_raw['Conceito'] = rmodelo_raw['Conceito'].replace('Fictício', 'TTI Total')    

    # Transformar o tipo das colunas SUBTIPO E LINHA para string da pivot_resultado
    pivot_resultado['SUBTIPO'] = pivot_resultado['SUBTIPO'].astype(str)
    pivot_resultado['LINHA'] = pivot_resultado['LINHA'].astype(str)

    # Separar os RModelos em total, contabil e ficticio
    rmodelo_total_raw = rmodelo_raw.iloc[:3646]
    rmodelo_contabil_raw = rmodelo_raw.iloc[3646:7090]
    rmodelo_ficticio_raw = rmodelo_raw.iloc[7090:] 

    # Criar dataframes e lista de dataframes em branco para preenchimento
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    df_list1 = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]

    df_list2 = deepcopy(df_list1)
    df_list3 = deepcopy(df_list1)

    df_list_final = []

    for i in range(len(list(segmentos.keys()))):
        dftotal_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Resultado') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[5]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list1[i] = pd.merge(rmodelo_total_raw, dftotal_filt, left_on = ['Cod'], right_on = ['LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfcontabil_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Resultado') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[6]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list2[i] = pd.merge(rmodelo_contabil_raw, dfcontabil_filt, left_on = ['Cod'], right_on = ['LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfficticio_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Resultado') & (pivot_resultado['SUBTIPO'] == subtipos[7]) & (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]
        df_list3[i] = pd.merge(rmodelo_ficticio_raw, dfficticio_filt, left_on = ['Cod'], right_on = ['LINHA'], how = 'left')

    for i in range(len(df_list1)):
        df_concat = pd.concat([df_list1[i], df_list2[i], df_list3[i]])
        df_list_final.append(df_concat)
        df_list_final[i] = df_list_final[i].drop(['HISTORICO_BP_MIS','SUBTIPO','SEGMENTO','LINHA'], axis=1) 

    col_list1= [f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12']
    col_list2= [f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12']
    col_list3= [f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12']
    col_list4= [f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12']
    col_list5= [f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12']

    if num_anos == 2:
        for i in range(len(df_list_final)):        
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].sum(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)
            
    elif num_anos == 3:
        for i in range(len(df_list_final)):        
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].sum(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)
            df_list_final[i][f'{y01}'] = df_list_final[i][col_list3].sum(axis=1)
    else:
        for i in range(len(df_list_final)):
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].sum(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)
            df_list_final[i][f'{y01}'] = df_list_final[i][col_list3].sum(axis=1)
            df_list_final[i][f'{y02}'] = df_list_final[i][col_list4].sum(axis=1)
            df_list_final[i][f'{y03}'] = df_list_final[i][col_list5].sum(axis=1)

    for i in range(len(df_list_final)):
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].astype(str)
        df_list_final[i] = df_list_final[i].sort_values(by=['ID'], ascending=True)
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('nan', '')

    return ajustar_row_cod_0(df_list_final)

def gerar_tabela_pmodelo(pmodelo_raw, pivot_resultado, subtipos, segmentos, num_anos, cenario):    
    '''Constrói o PModelo do Varejo Atual com Comercial, Moroso e Total.
    
    Parameters:
    pmodelo_raw(Pandas DataFrame): Template do PModelo com a estrutura (arquivo em Excel Modelo BP Atual)
    pivot_resultado(Pandas DataFrame): Tabela dinâmica oriunda da função gerar_pivot_table()
    subtipos(list): Subtipos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    segmentos(dict): Segmentos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    
    Returns:
    df_list_final(pandas DataFrame): PModelo final preenchido
    '''    
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    print(y0, y1, y01, y02, y02, y03)

    # Pegar apenas as linhas e colunas necessárias para utilizar no preenchimento do PModelo
    pmodelo_raw = pmodelo_raw.iloc[1:]
    pmodelo_raw = pmodelo_raw.drop(pmodelo_raw.iloc[:, 10:46], axis = 1)
    pmodelo_raw.columns = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada', 'Cod','Itens / Período','Novo Modelo', '', ' ', 'ISS', 'PIS', 'IR']    
    pmodelo_raw.dropna(subset=['Cod'])
    pmodelo_raw['ID'] = pmodelo_raw.index
    
    # Transformar o tipo das colunas Conceito e Cod para string do PModelo, além de substituir algumas strings para os joins
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].astype(str)
    pmodelo_raw['Cod'] = pmodelo_raw['Cod'].astype(str)
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].replace('Moroso', 'Morosa Total')
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].replace('Total', 'Carteira total')
    
    # Transformar o tipo das colunas SUBTIPO E LINHA para string da pivot_resultado
    pivot_resultado['SUBTIPO'] = pivot_resultado['SUBTIPO'].astype(str)
    pivot_resultado['LINHA'] = pivot_resultado['LINHA'].astype(str)

     # Separar os PModelos em Comercial, Moroso e Total
    pmodelo_comerc_raw = pmodelo_raw.loc[(pmodelo_raw['Conceito'] != 'Outros') & (pmodelo_raw['Conceito'] != 'Acordos') & (pmodelo_raw['Conceito'] != 'Morosa Total') & (pmodelo_raw['Conceito'] != 'Carteira total')]
    pmodelo_outros_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Outros']
    pmodelo_acordos_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Acordos']
    pmodelo_moroso_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Morosa Total']
    pmodelo_total_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Carteira total']   
    
    
    # Criar dataframes e lista de dataframes em branco para preenchimento
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    df_list1 = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]
    
    df_list2 = deepcopy(df_list1)
    df_list3 = deepcopy(df_list1)
    df_list4 = deepcopy(df_list1)
    df_list5 = deepcopy(df_list1)
    
    df_list_final = []
    

    # Preenchimento dos PModelos
    for i in range(len(list(segmentos.keys()))):
        dfcomerc_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo médio') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[0]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list1[i] = pd.merge(pmodelo_comerc_raw, dfcomerc_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfmorosa_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo médio') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[1]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list2[i] = pd.merge(pmodelo_moroso_raw, dfmorosa_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfcart_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo médio') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[2]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list3[i] = pd.merge(pmodelo_total_raw, dfcart_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfout_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo médio') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[3]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list4[i] = pd.merge(pmodelo_outros_raw, dfout_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfacor_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo médio') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[4]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list5[i] = pd.merge(pmodelo_acordos_raw, dfacor_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    # Empilhar os 3 tipos de PModelo
    for i in range(len(df_list1)):
        df_concat = pd.concat([df_list1[i], df_list4[i], df_list5[i], df_list2[i], df_list3[i]])
        df_list_final.append(df_concat)
        df_list_final[i] = df_list_final[i].drop(['HISTORICO_BP_MIS','SUBTIPO','SEGMENTO','LINHA'], axis=1) 
        
    # Gerar colunas ano consolidado        
    col_list1= [f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12']
    col_list2= [f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12']
    col_list3= [f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12']
    col_list4= [f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12']
    col_list5= [f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12']
    
    ultima_cols = [f'{y0}12',f'{y0}11',f'{y0}10',f'{y0}09',f'{y0}08',f'{y0}07',f'{y0}06',f'{y0}05',f'{y0}04',f'{y0}03',f'{y0}02',f'{y0}01']
    df_verifica_ultima_coluna = df_list_final[0].iloc[0:5]
    contador_lkw = 12
    for j in ultima_cols:
        if float(df_verifica_ultima_coluna[j].loc[df_verifica_ultima_coluna.Cod == '12']) != 0:
            break
        else:
            contador_lkw -= 1
            pass     

    if num_anos == 2:
        for i in range(len(df_list_final)):        
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].mean(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)/contador_lkw
            
    elif num_anos == 3:
        for i in range(len(df_list_final)):        
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].mean(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)/contador_lkw
            df_list_final[i][f'{y01}'] = df_list_final[i][col_list3].mean(axis=1)
            
    else:
        for i in range(len(df_list_final)):        
            df_list_final[i][f'{y1}'] = df_list_final[i][col_list1].mean(axis=1)
            df_list_final[i][f'{y0}'] = df_list_final[i][col_list2].sum(axis=1)/contador_lkw
            df_list_final[i][f'{y01}'] = df_list_final[i][col_list3].mean(axis=1)
            df_list_final[i][f'{y02}'] = df_list_final[i][col_list4].mean(axis=1)
            df_list_final[i][f'{y03}'] = df_list_final[i][col_list5].mean(axis=1)

    for i in range(len(df_list_final)):
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].astype(str)
        df_list_final[i] = df_list_final[i].sort_values(by=['ID'], ascending=True)
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('nan', '')
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('Morosa Total', 'Moroso')
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('Carteira total', 'Total')
        
    return ajustar_row_cod_0(df_list_final)    

def gerar_tabela_pmodelo_ponta(pmodelo_raw, pivot_resultado, subtipos, segmentos, num_anos, cenario):
    '''Constrói o PModelo Ponta do Varejo Atual com Comercial, Moroso e Total.
    
    Parameters:
    pmodelo_raw(Pandas DataFrame): Template do PModelo Ponta com a estrutura (arquivo em Excel Modelo BP Atual)
    pivot_resultado(Pandas DataFrame): Tabela dinâmica oriunda da função gerar_pivot_table()
    subtipos(list): Subtipos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    segmentos(dict): Segmentos para filtrar a pivot_resultado e preencher o rmodelo, oriundo do arquivo de variáveis (variaveis_var_atual)
    
    Returns:
    df_list_final(pandas DataFrame): PModelo final preenchido
    '''    
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    print(y0, y1, y01, y02, y02, y03)
    
    # Pegar apenas as linhas e colunas necessárias para utilizar no preenchimento do PModelo
    pmodelo_raw = pmodelo_raw.iloc[1:]
    pmodelo_raw = pmodelo_raw.drop(pmodelo_raw.iloc[:, 10:46], axis = 1)
    pmodelo_raw.columns = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada', 'Cod','Itens / Período','Novo Modelo', '', ' ', 'ISS', 'PIS', 'IR']  
    pmodelo_raw.dropna(subset=['Cod'])
    pmodelo_raw['ID'] = pmodelo_raw.index
    
    # Transformar o tipo das colunas Conceito e Cod para string do PModelo, além de substituir algumas strings para os joins
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].astype(str)
    pmodelo_raw['Cod'] = pmodelo_raw['Cod'].astype(str)
    
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].replace('Moroso', 'Morosa Total')
    pmodelo_raw['Conceito'] = pmodelo_raw['Conceito'].replace('Total', 'Carteira total')
    
    # Transformar o tipo das colunas SUBTIPO E LINHA para string da pivot_resultado
    pivot_resultado['SUBTIPO'] = pivot_resultado['SUBTIPO'].astype(str)
    pivot_resultado['LINHA'] = pivot_resultado['LINHA'].astype(str)

    # Separar os PModelos em Comercial, Moroso e Total
    pmodelop_comerc_raw = pmodelo_raw.loc[(pmodelo_raw['Conceito'] != 'Outros') & (pmodelo_raw['Conceito'] != 'Acordos') & (pmodelo_raw['Conceito'] != 'Morosa Total') & (pmodelo_raw['Conceito'] != 'Carteira total')]
    pmodelop_outros_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Outros']
    pmodelop_acordos_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Acordos']
    pmodelop_moroso_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Morosa Total']
    pmodelop_total_raw = pmodelo_raw.loc[pmodelo_raw['Conceito'] == 'Carteira total']   
    
    # Criar dataframes e lista de dataframes em branco para preenchimento
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    df_list1 = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]
    
    df_list2 = deepcopy(df_list1)
    df_list3 = deepcopy(df_list1)
    df_list4 = deepcopy(df_list1)
    df_list5 = deepcopy(df_list1)
    
    df_list_final = []

    # Preenchimento dos PModelos Ponta
    for i in range(len(list(segmentos.keys()))):
        dfcomerc_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo final') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[0]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list1[i] = pd.merge(pmodelop_comerc_raw, dfcomerc_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfmorosa_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo final') & 
                    (pivot_resultado['SUBTIPO'] == subtipos[1]) & 
                    (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list2[i] = pd.merge(pmodelop_moroso_raw, dfmorosa_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfcart_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo final') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[2]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list3[i] = pd.merge(pmodelop_total_raw, dfcart_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfout_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo final') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[3]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list4[i] = pd.merge(pmodelop_outros_raw, dfout_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    for i in range(len(list(segmentos.keys()))):
        dfacor_filt = pivot_resultado.loc[(pivot_resultado['HISTORICO_BP_MIS'] == 'Saldo final') & 
                     (pivot_resultado['SUBTIPO'] == subtipos[4]) & 
                     (pivot_resultado['SEGMENTO'] == list(segmentos.keys())[i])]        
        df_list5[i] = pd.merge(pmodelop_acordos_raw, dfacor_filt, left_on = ['Conceito','Cod'], right_on = ['SUBTIPO','LINHA'], how = 'left')

    # Empilhar os 3 tipos de PModelo
    for i in range(len(df_list1)):
        df_concat = pd.concat([df_list1[i], df_list4[i], df_list5[i], df_list2[i], df_list3[i]])
        df_list_final.append(df_concat)
        df_list_final[i] = df_list_final[i].drop(['HISTORICO_BP_MIS','SUBTIPO','SEGMENTO','LINHA'], axis=1) 
        
    
    # Gerar colunas de ano consolidados
    ultima_cols = [f'{y0}12',f'{y0}11',f'{y0}10',f'{y0}09',f'{y0}08',f'{y0}07',f'{y0}06',f'{y0}05',f'{y0}04',f'{y0}03',f'{y0}02',f'{y0}01']
    df_verifica_ultima_coluna = df_list_final[0].iloc[0:5]
    
    
    if num_anos == 2:
        for i in range(len(df_list_final)):
            df_list_final[i][f'{y1}'] = df_list_final[i][f'{y1}12']

            for j in ultima_cols:
                if float(df_verifica_ultima_coluna[j].loc[df_verifica_ultima_coluna.Cod == '12']) != 0:
                    df_list_final[i][f'{y0}'] = df_list_final[i][j]
                    break
                else:
                    pass  
                
    elif num_anos == 3:
        for i in range(len(df_list_final)):
            df_list_final[i][f'{y1}'] = df_list_final[i][f'{y1}12']
            df_list_final[i][f'{y01}'] = df_list_final[i][f'{y01}12']

            for j in ultima_cols:
                if float(df_verifica_ultima_coluna[j].loc[df_verifica_ultima_coluna.Cod == '12']) != 0:
                    df_list_final[i][f'{y0}'] = df_list_final[i][j]
                    break
                else:
                    pass  
                
                
    else:
        for i in range(len(df_list_final)):
            df_list_final[i][f'{y1}'] = df_list_final[i][f'{y1}12']
            df_list_final[i][f'{y01}'] = df_list_final[i][f'{y01}12']
            df_list_final[i][f'{y02}'] = df_list_final[i][f'{y02}12']
            df_list_final[i][f'{y03}'] = df_list_final[i][f'{y03}12']

            for j in ultima_cols:
                if float(df_verifica_ultima_coluna[j].loc[df_verifica_ultima_coluna.Cod == '12']) != 0:
                    df_list_final[i][f'{y0}'] = df_list_final[i][j]
                    break
                else:
                    pass  

    for i in range(len(df_list_final)):
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].astype(str)
        df_list_final[i] = df_list_final[i].sort_values(by=['ID'], ascending=True)
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('nan', '')
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('Morosa Total', 'Moroso')
        df_list_final[i]['Conceito'] = df_list_final[i]['Conceito'].replace('Carteira total', 'Total')
        
    return ajustar_row_cod_0(df_list_final)

def gerar_tabela_smodelo(pmodelo, rmodelo, smodelo_raw, dic_dias_mes, dic_du_mes, anos, num_anos, cenario):
    '''Constrói o SModelo do Varejo Atual.
    
    Parameters:
    pmodelo(Pandas DataFrame): PModelo final preenchido
    rmodelo(Pandas DataFrame): RModelo final preenchido
    smodelo_raw(Pandas DataFrame): Template do SModelo com a estrutura (arquivo em Excel Modelo BP Atual)
    dic_dias_mes(dict): Meses e Anos e seus números de dias corridos, oriundo do arquivo de variáveis (variaveis_var_atual)
    dic_du_mes(dict): Meses e Anos e seus números de dias úteis, oriundo do arquivo de variáveis (variaveis_var_atual)
    anos(str): Número de anos para o cálculo do spread (2, 3 ou 5). Os demais anos do P serão preenchidos com 0 para melhor performance
    
    Returns:
    smodelo_final(pandas DataFrame): SModelo final preenchido
    '''  
    # Pegar apenas as linhas e colunas necessárias para utilizar no preenchimento do SModelo de todas as linhas
    s_modelo_padrao = smodelo_raw.iloc[1:]
    s_modelo_padrao = s_modelo_padrao.drop(s_modelo_padrao.iloc[:, 9:40], axis = 1)
    s_modelo_padrao = s_modelo_padrao.drop(s_modelo_padrao.iloc[:, 8:10], axis = 1)
    s_modelo_padrao = s_modelo_padrao.drop(s_modelo_padrao.iloc[:, 9:13], axis = 1)
    s_modelo_padrao.columns = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada', 'Cod', 'Itens / Período', 'Novo Modelo', 'Param', 'ISS', 'PIS', 'IR']
    s_modelo_padrao['Cod'] = s_modelo_padrao['Cod'].astype(str)
    s_modelo_padrao['ID'] = s_modelo_padrao.index

    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    print(y0, y1, y01, y02, y02, y03)

    
    if num_anos == 2:
        cols_val = ['ID','Cod', f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}']
        
    elif num_anos == 3:
        cols_val = ['ID','Cod', f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                    f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}']
        
    else:
        cols_val = ['ID','Cod', f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                    f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                    f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                    f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}']

    #Criar cópias do PModelo e RModelo para uso posterior
    s_modelo_raw_p = deepcopy(pmodelo)
    s_modelo_raw_rtot = deepcopy(rmodelo)

    for i in range(len(s_modelo_raw_p)):
        s_modelo_raw_p[i] = s_modelo_raw_p[i][cols_val]
        s_modelo_raw_rtot[i] = s_modelo_raw_rtot[i][cols_val]


        #Criar Dataframes vazios e colocá-los em listas de DataFrames
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    s_modelo_raw_r = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]

    smodelo_ant = smodelo = smodelo_final = deepcopy(s_modelo_raw_r)

    # Filtrar o RModelo para pegar somente o Total
    for i in range(len(s_modelo_raw_rtot)):
        s_modelo_raw_r[i] = s_modelo_raw_rtot[i].iloc[:3000]

    for i in range(len(s_modelo_raw_p)):
        s_modelo_raw_p[i]['Cod'] = s_modelo_raw_p[i]['Cod'].astype(str)
        s_modelo_raw_r[i]['Cod'] = s_modelo_raw_r[i]['Cod'].astype(str)


    # Pegar apenas as linhas e colunas necessárias para utilizar no preenchimento do SModelo excluindo as linhas com Cod 'x', 0 e 'nan'
    smodelo_raw = smodelo_raw.iloc[1:]
    smodelo_raw = smodelo_raw.drop(smodelo_raw.iloc[:, 0:5], axis = 1)
    smodelo_raw = smodelo_raw.drop(smodelo_raw.iloc[:, 1:36], axis = 1)
    smodelo_raw = smodelo_raw.drop(smodelo_raw.iloc[:, 2:], axis = 1)
    smodelo_raw.columns = ['Cod','Param']
    smodelo_raw.drop(smodelo_raw[smodelo_raw['Cod'] == 'x'].index, inplace = True) 
    smodelo_raw.drop(smodelo_raw[smodelo_raw['Cod'] == 0].index, inplace = True) 
    smodelo_raw.drop(smodelo_raw[smodelo_raw['Cod'] == 'nan'].index, inplace = True) 
    smodelo_raw['Cod'] = smodelo_raw['Cod'].astype(str)


    #Manter somente as linhas presentes no SModelo, desconsiderando o restante que está no PModelo, mas não no SModelo
    for i in range(len(s_modelo_raw_p)):
        s_modelo_raw_p[i] = s_modelo_raw_p[i].iloc[:1000]

    #Remover linhas em branco dos DataFrames   
    for j in range(len(s_modelo_raw_p)):
        s_modelo_raw_p[j].drop(s_modelo_raw_p[j][s_modelo_raw_p[j]['Cod'] == '0'].index, inplace = True)
        s_modelo_raw_p[j].drop(s_modelo_raw_p[j][s_modelo_raw_p[j]['Cod'] == 'x'].index, inplace = True)


    #Remover linhas em branco dos DataFrames
    for j in range(len(s_modelo_raw_r)):
        s_modelo_raw_r[j].drop(s_modelo_raw_r[j][s_modelo_raw_r[j]['Cod'] == '0'].index, inplace = True)
        s_modelo_raw_r[j].drop(s_modelo_raw_r[j][s_modelo_raw_r[j]['Cod'] == 'x'].index, inplace = True)
        s_modelo_raw_r[j].drop(s_modelo_raw_r[j][s_modelo_raw_r[j]['Cod'] == 'nan'].index, inplace = True)

    #Realizar left join entre os Dataframes de PModelo e o SModelo, para trazer a coluna de parâmetro de cálculo para Spread (se dia útil ou não) ao PModelo
    for i in range(len(s_modelo_raw_p)):
        smodelo_ant[i] = pd.merge(s_modelo_raw_p[i], smodelo_raw, on = ['Cod'], how = 'left')

    #Realizar left join para trazer aos Dataframes acima, as colunas de valor do RModelo
    for i in range(len(smodelo_ant)):
        smodelo[i] = pd.merge(smodelo_ant[i], s_modelo_raw_r[i], on = ['Cod'], how = 'left')


    #Renomear colunas: _x são colunas do PModelo, _y do RModelo
    mes = ['01','02','03','04','05','06','07','08','09','10','11','12']
    
    if num_anos == 2:
        ano = [y1, y0]
    elif num_anos == 3:
        ano = [y1, y0, y01]
    else:
        ano = [y1, y0, y01, y02, y03]
    
    for i in range(len(smodelo)):
        for a in ano:
            for j in mes:
                smodelo[i] = smodelo[i].rename(columns={f'{a}{j}_x' : f'{a}{j}_p',
                                                        f'{a}_x' : f'{a}_p',
                                                        f'{a}{j}_y' : f'{a}{j}_r',
                                                        f'{a}_y' : f'{a}_r'})
        
    cols_numericas = ['Param']
    
    #Construção de listas para posterior cálculo do spread
    for a in ano:
        for j in mes:
            col1 = f'{a}{j}_p'
            col2 = f'{a}_p'
            col3 = f'{a}{j}_r'
            col4 = f'{a}_r'
            cols_numericas.append(col1)
            cols_numericas.append(col2)
            cols_numericas.append(col3)
            cols_numericas.append(col4)

    
    #Substituir espaço vazio po NaN
    for i in range(len(smodelo)):
        for j in cols_numericas:
            smodelo[i][j] = smodelo[i][j].replace('',np.nan)
    
    #Calcular Colunas Spread
    if anos == '2' and num_anos == 5:
        for i in range(len(smodelo)):
            for a in ano:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}{j}_r', f'{a}{j}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}{j}', f'{a}'), axis = 1)
                    smodelo[i][f'{a}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}_r', f'{a}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}', f'{a}'), axis = 1)
        for i in range(len(smodelo)):
            for a in ano[2:]:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = 0
                    smodelo[i][f'{a}'] = 0
                    
                    
    elif anos == '3' and num_anos == 5:
        for i in range(len(smodelo)):
            for a in ano:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}{j}_r', f'{a}{j}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}{j}', f'{a}'), axis = 1)
                    smodelo[i][f'{a}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}_r', f'{a}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}', f'{a}'), axis = 1)
        for i in range(len(smodelo)):
            for a in ano:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = 0
                    smodelo[i][f'{a}'] = 0
                    
                    
    elif anos == '5' and num_anos == 5:
        for i in range(len(smodelo)):
            for a in ano:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}{j}_r', f'{a}{j}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}{j}', f'{a}'), axis = 1)
                    smodelo[i][f'{a}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}_r', f'{a}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}', f'{a}'), axis = 1)
                    
    elif anos in ('2', '3') and num_anos != 5:
        for i in range(len(smodelo)):
            for a in ano:
                for j in mes:
                    smodelo[i][f'{a}{j}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}{j}_r', f'{a}{j}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}{j}', f'{a}'), axis = 1)
                    smodelo[i][f'{a}'] = smodelo[i].apply(lambda row: tratar_cols_spread(row, f'{a}_r', f'{a}_p', 'Param', dic_dias_mes, dic_du_mes, f'{a}', f'{a}'), axis = 1)


    # Selecionar colunas que serão usadas
    if num_anos == 2:
        for i in range(len(smodelo)):
            smodelo[i] = smodelo[i][['ID_x','Cod',f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}']]
            
    elif num_anos == 3:
        for i in range(len(smodelo)):
            smodelo[i] = smodelo[i][['ID_x','Cod',f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                    f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}']]
            
    else:
        for i in range(len(smodelo)):
            smodelo[i] = smodelo[i][['ID_x','Cod',f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                    f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                    f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                                    f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                                    f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}']]

    #Renomear coluna ID
    for i in range(len(smodelo)):
        smodelo[i] = smodelo[i].rename(columns={'ID_x':'ID'})

    # Left join entre smodelo e smodelo_padrao
    for i in range(len(smodelo)):
        smodelo_final[i] = pd.merge(s_modelo_padrao, smodelo[i], on = ['ID'], how = 'left')

    # Pegar apenas as colunas necessárias na ordem necessária
    if num_anos == 2:
        for i in range(len(smodelo_final)):
            smodelo_final[i] = smodelo_final[i][['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada','Cod_x', 'Itens / Período', 'Novo Modelo',
                                                 f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                                 f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                                 'Param', 'ISS', 'PIS', 'IR']]
    elif num_anos == 3:
        for i in range(len(smodelo_final)):
            smodelo_final[i] = smodelo_final[i][['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada','Cod_x', 'Itens / Período', 'Novo Modelo',
                                                 f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                                 f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                                 f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                                                 'Param', 'ISS', 'PIS', 'IR']]
    else:
        for i in range(len(smodelo_final)):
            smodelo_final[i] = smodelo_final[i][['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cascada','Cod_x', 'Itens / Período', 'Novo Modelo',
                                                 f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                                 f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                                 f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                                                 f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                                                 f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                                                 'Param', 'ISS', 'PIS', 'IR']]

    # Renomear colunas
    for i in range(len(smodelo_final)):
        smodelo_final[i] = smodelo_final[i].rename(columns={'Cod_x': 'Cod',
                                                            'Param': 'FL_DIAS_CORRIDOS'})
        
    return ajustar_row_cod_0(smodelo_final)

def gerar_tabela_cascmodelo(cmodelo_raw, rmodelo_final, num_anos, cenario):
    '''Constrói o CascModelo do Varejo Atual.
    
    Parameters:
    cmodelo_raw(Pandas DataFrame): Template do CModelo com a estrutura (arquivo em Excel Modelo BP Atual)
    rmodelo_final(Pandas DataFrame): RModelo final preenchido
    
    Returns:
    cmodelo_geral(pandas DataFrame): CModelo final preenchido
    '''  
     # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    print(y0, y1, y01, y02, y02, y03)

    # Pular 1ª linha em branco
    cmodelo_raw = cmodelo_raw.iloc[1:]

    # Renomear coluna
    cmodelo_raw = cmodelo_raw.rename(columns={'Unnamed: 5': 'Cod'}) 

    # Criar coluna 'ID' para juntar as partes do dataset mais tarde na ordem correta
    cmodelo_raw['ID'] = cmodelo_raw.index

    # Lista de Cod na Cmodelo cujo join será feito com a Chave Cascada do Rmodelo
    cod_chave_casc = [13, 14, 15, 16, 17, 18, 19, 20, 22, 24, 25, 26, 27, 28, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 59, 60, 61, 62, 63, 64, 65]

    # Cria dataframe somente com os Cod da lista cod_chave_casc
    cmodelo_raw_chave_casc = cmodelo_raw.loc[cmodelo_raw['Cod'].isin(cod_chave_casc)]
    cmodelo_raw_chave_casc['Chave Cascada'] = cmodelo_raw_chave_casc.apply(lambda row: tratar_col_casc(row, 'Itens / Período', 'Cod'),
                                                                                 axis = 1)

    # Cria dataframe somente com os Cod fora da lista cod_chave_casc
    cmodelo_raw_casc = cmodelo_raw.loc[~cmodelo_raw['Cod'].isin(cod_chave_casc)]


    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    rmodelo_chave_casc_agr = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]

    rmodelo_casc = deepcopy(rmodelo_chave_casc_agr)
    rmodelo_chave_casc = deepcopy(rmodelo_chave_casc_agr)
    rmodelo_casc_agr = deepcopy(rmodelo_chave_casc_agr)
    cmodelo_chave_casc = deepcopy(rmodelo_chave_casc_agr)
    cmodelo_casc = deepcopy(rmodelo_chave_casc_agr)
    rmodelo_tot = deepcopy(rmodelo_chave_casc_agr)
    cmodelo_geral = deepcopy(rmodelo_chave_casc_agr)
    
    
    rmodelo = deepcopy(rmodelo_final)
    for i in range(len(rmodelo)):
        rmodelo[i] = rmodelo[i].iloc[:2925]
        
    rmodelo_casc = deepcopy(rmodelo)
    rmodelo_chave_casc = deepcopy(rmodelo)


    for i in range(len(rmodelo_casc)):
        rmodelo_casc[i].drop(rmodelo_casc[i][rmodelo_casc[i]['Cod'] == '0'].index, inplace = True)
        rmodelo_casc[i].drop(rmodelo_casc[i][rmodelo_casc[i]['Cod'] == 'x'].index, inplace = True)
        rmodelo_casc[i].drop(rmodelo_casc[i][rmodelo_casc[i]['Cod'] == 'nan'].index, inplace = True)
        rmodelo_casc[i].drop(['Totalizador', 'Chave Cascada', 'Alíquotas', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos', 'Linha excel',
                              f'Reparto % {y1}', f'Reparto % {y0}', 'ISS', 'PIS', 'IR',''],
                                 axis = 1, inplace = True)
        rmodelo_casc_agr[i] = rmodelo_casc[i].groupby(['Cascada']).sum().reset_index()
        cmodelo_casc[i] = pd.merge(cmodelo_raw_casc, rmodelo_casc_agr[i], left_on = ['Itens / Período'], right_on = ['Cascada'], how = 'left')
        cmodelo_casc[i].drop(['Cascada', 'ID_y'], axis = 1, inplace = True)

    for i in range(len(rmodelo_chave_casc)):
        rmodelo_chave_casc[i].drop(rmodelo_chave_casc[i][rmodelo_chave_casc[i]['Cod'] == '0'].index, inplace = True)
        rmodelo_chave_casc[i].drop(rmodelo_chave_casc[i][rmodelo_chave_casc[i]['Cod'] == 'x'].index, inplace = True)
        rmodelo_chave_casc[i].drop(rmodelo_chave_casc[i][rmodelo_chave_casc[i]['Cod'] == 'nan'].index, inplace = True)
        rmodelo_chave_casc[i].drop(['Totalizador', 'Cascada', 'Alíquotas', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos', 'Linha excel',
                                   f'Reparto % {y1}', f'Reparto % {y0}', 'ISS', 'PIS', 'IR',''],
                                   axis = 1, inplace = True)
        rmodelo_chave_casc_agr[i] = rmodelo_chave_casc[i].groupby(['Chave Cascada']).sum().reset_index()
        cmodelo_chave_casc[i] = pd.merge(cmodelo_raw_chave_casc, rmodelo_chave_casc_agr[i], on = ['Chave Cascada'], how = 'left')
        cmodelo_chave_casc[i].drop(['Chave Cascada', 'ID_y'], axis = 1, inplace = True)

    for i in range(len(cmodelo_geral)):
        cmodelo_geral[i] = pd.concat([cmodelo_casc[i], cmodelo_chave_casc[i]])

    for i in range(len(cmodelo_geral)):
        cmodelo_geral[i] = cmodelo_geral[i].sort_values(by=['ID_x'], ascending=True)

    cols = cmodelo_geral[0].columns
    cols = cols.drop(['Cod', 'Itens / Período', 'ID_x'])

    for i in range(len(cmodelo_geral)):
        for j in cols:
        #Gerar colunas para a linha Margem c/ Produtos
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 12] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 13) & (cmodelo_geral[i]['Cod'] <= 28)].sum()

        #Gerar colunas para a linha Provisiones de Insolvencias
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 29] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 30) & (cmodelo_geral[i]['Cod'] <= 33)].sum()

        #Gerar colunas para a linha Alocação de Capital + Remun. Resultado
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 34] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 35) & (cmodelo_geral[i]['Cod'] <=37)].sum()

        #Gerar colunas para a linha Margem de Intermediação (líquida)
            s1 = 0
            values1 = []

            aloc_cap1 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 34])
            prov_ins1 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 29])
            marg_prod1 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 12])
            values1.append(aloc_cap1)
            values1.append(prov_ins1)
            values1.append(marg_prod1)

            for u1 in values1:
                if pd.isna(u1) == True:
                    pass
                else:
                    s1 += u1      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 38] = s1

            #Gerar colunas para a linha Comissões
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 40] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 41) & (cmodelo_geral[i]['Cod'] <= 65)].sum()

            #Gerar colunas para a linha Margem Basica Líquida
            s2 = 0
            values2 = []

            marg_int_liq2 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 38])
            comissoes2 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 40])
            values2.append(marg_int_liq2)
            values2.append(comissoes2)

            for u2 in values2:
                if pd.isna(u2) == True:
                    pass
                else:
                    s2 += u2      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 67] = s2

            #Gerar colunas para a linha Margem Basica Bruta
            s3 = 0
            values3 = []

            marg_bas_liq3 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 67])
            prov_ins3 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 29])
            values3.append(marg_bas_liq3)
            values3.append(prov_ins3)

            s3 = values3[0] - values3[1]

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 66] = s3

            #Gerar colunas para a linha ROF
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 72] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 73) & (cmodelo_geral[i]['Cod'] <= 82)].sum()

            #Gerar colunas para a linha Outros Resultados Operacionais
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 83] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 84) & (cmodelo_geral[i]['Cod'] <= 90)].sum()

            #Gerar colunas para a linha Margem Ordinária Bruta
            s4 = 0
            values4 = []

            marg_bas_brut4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 66])
            rof4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 72])
            out_res_oper4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 83])
            equiv4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 69])
            div4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 70])
            lin_s_inf4 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 71])
            
            values4.append(marg_bas_brut4)
            values4.append(rof4)
            values4.append(out_res_oper4)
            values4.append(equiv4)
            values4.append(div4)
            values4.append(lin_s_inf4)

            for u4 in values4:
                if pd.isna(u4) == True:
                    pass
                else:
                    s4 += u4      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 92] = s4

            #Gerar colunas para a linha Margem Ordinária Líquida
            s5 = 0
            values5 = []

            marg_bas_liq5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 67])
            rof5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 72])
            out_res_oper5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 83])
            equiv5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 69])
            div5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 70])
            lin_s_inf5 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 71])
            values5.append(marg_bas_liq5)
            values5.append(rof5)
            values5.append(out_res_oper5)
            values5.append(equiv5)
            values5.append(div5)
            values5.append(lin_s_inf5)

            for u5 in values5:
                if pd.isna(u5) == True:
                    pass
                else:
                    s5 += u5      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 93] = s5

            #Gerar colunas para a linha xxxxLinha s/ Informação
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 95] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 96) & (cmodelo_geral[i]['Cod'] <= 101)].sum()

            #Gerar colunas para a linha Gastos Gerais da Administração
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 102] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 103) & (cmodelo_geral[i]['Cod'] <= 106)].sum()

            #Gerar colunas para a linha BAI - Lucro antes dos Impostos
            s6 = 0
            values6 = []

            marg_ord_liq6 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 93])
            lin_s_inf6 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 95])
            gast_ger_adm6 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 102])
            oryp6 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 107])
            out_ativ6 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 108])

            values6.append(marg_ord_liq6)
            values6.append(lin_s_inf6)
            values6.append(gast_ger_adm6)
            values6.append(oryp6)
            values6.append(out_ativ6)


            for u6 in values6:
                if pd.isna(u6) == True:
                    pass
                else:
                    s6 += u6      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 111] = s6

            #Gerar colunas para a linha Benefício antes do IR
            s7 = 0
            values7 = []

            bai7 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 111])
            iss7 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 113])
            pis_cofins7 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 114])

            values7.append(bai7)
            values7.append(iss7)
            values7.append(pis_cofins7)


            for u7 in values7:
                if pd.isna(u7) == True:
                    pass
                else:
                    s7 += u7      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 115] = s7

            #Gerar colunas para a linha BDI - Lucro Líquido
            s8 = 0
            values8 = []

            benef_ant_ir8 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 115])
            ir_cs8 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 116])

            values8.append(benef_ant_ir8)
            values8.append(ir_cs8)

            for u8 in values8:
                if pd.isna(u8) == True:
                    pass
                else:
                    s8 += u8      

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 118] = s8

            #Gerar colunas para a linha (-) Minoritários
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 119] = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 118]) * 0.117

            #Gerar colunas para a linha Lucro Atribuído
            s9 = 0
            values9 = []

            bdi9 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 118])
            min9 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 119])
            values9.append(bdi9)
            values9.append(min9)

            s9 = values9[0] - values9[1]

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 120] = s9     

            #Gerar colunas para a linha Custo Demais Ativos
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 122] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1216'])

            #Gerar colunas para a linha Remuneração de Capital
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 123] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1217'])

            #Gerar colunas para a linha Remuneração de Resultado
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 124] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1218'])

            #Gerar colunas para a linha PIS / Cofins
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 125] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1219'])

            #Gerar colunas para a linha Imposto de Renda
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 126] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1220'])

            #Gerar colunas para a linha BDI critério Espanha
            bdi10 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 118])

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 127] = bdi10 + cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 122) & (cmodelo_geral[i]['Cod'] <= 126)].sum()

            #Gerar colunas para a linha (-) Minoritários critério Espanha 
            bdi_crit_esp11 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 127])

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 128] = bdi_crit_esp11 * 0.117

            #Gerar colunas para a linha Lucro Atribuído critério Espanha 
            s12 = 0
            values12 = []

            bdi_crit_esp12 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 127])
            min_crit_esp12 = float(cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 128])
            values12.append(bdi_crit_esp12)
            values12.append(min_crit_esp12)

            s12 = values12[0] - values12[1]

            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 129] = s12        

            #Gerar colunas para a linha Detalhe Custo Demais Ativos
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 130] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1262'])

            #Gerar colunas para a linha Custo Demais Ativos
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 132] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1264'])

            #Gerar colunas para a linha Sofware
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 133] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1265'])

            #Gerar colunas para a linha Hardware
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 134] = float(rmodelo[i][j].loc[rmodelo[i]['Cod'] == '1266'])

            #Gerar colunas para a linha Total Custos Demais Ativos
            cmodelo_geral[i][j].loc[cmodelo_geral[i]['Cod'] == 131] = cmodelo_geral[i][j].loc[(cmodelo_geral[i]['Cod'] >= 132) & (cmodelo_geral[i]['Cod'] <= 134)].sum()  
            
    #Gerar colunas anos consolidados
        
    col_list1= [f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12']
    col_list2= [f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12']
    col_list3= [f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12']
    col_list4= [f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12']
    col_list5= [f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12']

    if num_anos == 2:
        for i in range(len(cmodelo_geral)):
            cmodelo_geral[i][f'{y1}'] = cmodelo_geral[i][col_list1].sum(axis=1)
            cmodelo_geral[i][f'{y0}'] = cmodelo_geral[i][col_list2].sum(axis=1)
            
    elif num_anos == 3:
        for i in range(len(cmodelo_geral)):
            cmodelo_geral[i][f'{y1}'] = cmodelo_geral[i][col_list1].sum(axis=1)
            cmodelo_geral[i][f'{y0}'] = cmodelo_geral[i][col_list2].sum(axis=1)
            cmodelo_geral[i][f'{y01}'] = cmodelo_geral[i][col_list3].sum(axis=1)
            
    else:
        for i in range(len(cmodelo_geral)):
            cmodelo_geral[i][f'{y1}'] = cmodelo_geral[i][col_list1].sum(axis=1)
            cmodelo_geral[i][f'{y0}'] = cmodelo_geral[i][col_list2].sum(axis=1)
            cmodelo_geral[i][f'{y01}'] = cmodelo_geral[i][col_list3].sum(axis=1)
            cmodelo_geral[i][f'{y02}'] = cmodelo_geral[i][col_list4].sum(axis=1)
            cmodelo_geral[i][f'{y03}'] = cmodelo_geral[i][col_list5].sum(axis=1)
            
    return ajustar_row_cod_0(cmodelo_geral)

def ajustar_colunas_dfs(pmodelo, pmodelop, rmodelo, smodelo, cascmodelo, num_anos, cenario):
    '''Ajuste final na estrutura dos Modelos antes de passar para Excel.
    
    Parameters:
    pmodelo(Pandas DataFrame): PModelo final preenchido
    pmodelop(Pandas DataFrame): PModelo Ponta final preenchido
    rmodelo(Pandas DataFrame): RModelo final preenchido
    smodelo(Pandas DataFrame): SModelo final preenchido
    cascmodelo(Pandas DataFrame): CascModelo final preenchido
    
    Returns:
    pmodelo(Pandas DataFrame): PModelo final ajustado
    pmodelop(Pandas DataFrame): PModelo Ponta final ajustado
    rmodelo(Pandas DataFrame): RModelo final ajustado
    smodelo(Pandas DataFrame): SModelo final ajustado
    cascmodelo(Pandas DataFrame): CascModelo final ajustado
    '''  
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    
    # Seleciona quais colunas serão utilizadas e em qual ordem de cada Modelo
    if num_anos == 2:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', 
                     'ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'Reparto % {y0}', 'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                    'FL_DIAS_CORRIDOS','ISS','PIS','IR']

        ccols_val = ['Cod','Itens / Período',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}']
        
    elif num_anos == 3:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', 
                     'ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', f'Reparto % {y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                    'FL_DIAS_CORRIDOS','ISS','PIS','IR']

        ccols_val = ['Cod','Itens / Período',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}']
        
    else:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}', 
                     'ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', f'Reparto % {y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                     'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                    'FL_DIAS_CORRIDOS','ISS','PIS','IR']

        ccols_val = ['Cod','Itens / Período',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}']
    
    
    # Converter a coluna Cod de todos os modelos para inteiro
    for i in range(len(smodelo)):   
        smodelo[i]['Cod'] = smodelo[i]['Cod'].replace('x', -999)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].replace('x', -999)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].replace('x', -999)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].replace('x', -999)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].replace('x', -999)        
        smodelo[i]['Cod'] = smodelo[i]['Cod'].replace('', -999)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].replace('', -999)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].replace('', -999)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].replace('', -999)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].replace('', -999)
        smodelo[i]['Cod'] = smodelo[i]['Cod'].replace('nan', -999)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].replace('nan', -999)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].replace('nan', -999)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].replace('nan', -999)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].replace('nan', -999)
        smodelo[i]['Cod'] = smodelo[i]['Cod'].fillna(-999)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].fillna(-999)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].fillna(-999)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].fillna(-999)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].fillna(-999)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].replace('Linha', -999)
        
        smodelo[i]['Cod'] = smodelo[i]['Cod'].astype(int)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].astype(int)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].astype(int)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].astype(int)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].astype(int)
        
        
        smodelo[i]['Cod'] = smodelo[i]['Cod'].replace(-999, np.nan)
        pmodelo[i]['Cod'] = pmodelo[i]['Cod'].replace(-999, np.nan)
        pmodelop[i]['Cod'] = pmodelop[i]['Cod'].replace(-999, np.nan)
        rmodelo[i]['Cod'] = rmodelo[i]['Cod'].replace(-999, np.nan)
        cascmodelo[i]['Cod'] = cascmodelo[i]['Cod'].replace(-999, np.nan) 
       
    # Filtra os modelos pelas colunas designadas anteriormente
    for i in range(len(pmodelo)):
        pmodelo[i] = pmodelo[i][pcols_val]
        pmodelop[i] = pmodelop[i][pcols_val]
        rmodelo[i] = rmodelo[i][rcols_val]
        smodelo[i] = smodelo[i][scols_val]
        cascmodelo[i] = cascmodelo[i][ccols_val]
    
    return pmodelo, pmodelop, rmodelo, smodelo, cascmodelo

def gerar_headers_finais(tb_analitica, num_anos, cenario):
    '''Constrói os headers finais de cenários para os modelos do Varejo Atual
    
    Parameters:
    tb_analitica(Pandas DataFrame): Dataframe da TB_ANALITICA_PREVIA (arquivo em Excel)
    
    Returns:
    header_p(pandas DataFrame): Header final para os PModelos
    header_r(pandas DataFrame): Header final para o RModelo
    header_s(pandas DataFrame): Header final para o SModelo
    header_casc(pandas DataFrame): Header final para o Casc Modelo
    '''
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    
    #Criar lista de meses e multiplicá-la por 5 (para 5 anos) e utilizar em loop for
    mes = ['01','02','03','04','05','06','07','08','09','10','11','12']
    temp_mes = list(mes)
    count = 5
    for i in range(count):
        for element in temp_mes:
            mes.append(element)
            
    #Criar lista de anos com 12 posições para cada ano e utilizar em loop for
    if num_anos == 2:
        ano = [y1, y0]
        temp_ano = []
        temp_ano += 12*[y1]
        temp_ano += 12*[y0]
    elif num_anos == 3:
        ano = [y1, y0, y01]
        temp_ano = []
        temp_ano += 12*[y1]
        temp_ano += 12*[y0]
        temp_ano += 12*[y01]
    else:
        ano = [y1, y0, y01, y02, y03]
        temp_ano = []
        temp_ano += 12*[y1]
        temp_ano += 12*[y0]
        temp_ano += 12*[y01]
        temp_ano += 12*[y02]
        temp_ano += 12*[y03]
        

    
    header_p0 = tb_analitica[['MES','CENARIO']].drop_duplicates().sort_values(by=['MES']).reset_index(drop=True).T
    
    for (a,b,c) in zip(range(60), temp_ano, mes):
        header_p0 = header_p0.rename(columns={a: f'{b}{c}'})
        
    header_p0['Totalizador'] = 'Valores em R$ Mil'
    header_p0['Chave Cascada'] = ''*1
    header_p0['Alíquotas'] = ''*2
    header_p0['Cascada'] = ''*3
    header_p0['Conceito'] = ''*4
    header_p0['Cod'] = ''*5
    header_p0['Itens / Período'] = ''*6
    header_p0['Novo Modelo'] = ''*7
    header_p0[''] = ''*8
    header_p0[' '] = ''*9
    header_p0['ISS'] = ''*10
    header_p0['PIS'] = ''*11
    header_p0['IR'] = ''*12
    header_p0[f'{y1}'] = header_p0[f'{y1}12']
    header_p0[f'{y0}'] = header_p0[f'{y0}12']
    
    if num_anos == 3:
        header_p0[f'{y01}'] = header_p0[f'{y01}12']
    elif num_anos == 5:
        header_p0[f'{y01}'] = header_p0[f'{y01}12']
        header_p0[f'{y02}'] = header_p0[f'{y02}12']
        header_p0[f'{y03}'] = header_p0[f'{y03}12']
    
    if num_anos == 2:
        header_p0 = header_p0[['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', 
                               'ISS', 'PIS', 'IR']]
    elif num_anos == 3:
        header_p0 = header_p0[['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', 
                               'ISS', 'PIS', 'IR']]
    else:
        header_p0 = header_p0[['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '', ' ',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                               f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                               f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}', 
                               'ISS', 'PIS', 'IR']]


    header_p0 = header_p0.iloc[1:]
    header_p0.reset_index(drop = True, inplace=True)
    
    n = 9
    new_index = pd.RangeIndex(len(header_p0)*(n+1))
    header_p = pd.DataFrame(np.nan, index=new_index, columns=header_p0.columns)
    ids = np.arange(len(header_p0))*(n+1)
    header_p.loc[ids] = header_p0.values
    
    header_s0 = deepcopy(header_p0)
    header_s0['FL_DIAS_CORRIDOS'] = ''*13
    header_s0['Totalizador'] = ''*14
    
    if num_anos == 2:
        header_s0 = header_s0[['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                               'FL_DIAS_CORRIDOS','ISS','PIS','IR']]
    elif num_anos == 3:
        header_s0 = header_s0[['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                               'FL_DIAS_CORRIDOS','ISS','PIS','IR']]
    else:
        header_s0 = header_s0[['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo',
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                               f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                               f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                               'FL_DIAS_CORRIDOS','ISS','PIS','IR']]
    
    n = 9
    new_index = pd.RangeIndex(len(header_s0)*(n+1))
    header_s = pd.DataFrame(np.nan, index=new_index, columns=header_s0.columns)
    ids = np.arange(len(header_s0))*(n+1)
    header_s.loc[ids] = header_s0.values
    
    header_r0 = deepcopy(header_p0)
    header_r0['Segmentos'] = ''*15
    header_r0['Linha excel'] = ''*16
    header_r0[f'Reparto % {y1}'] = ''*17
    header_r0[f'Reparto % {y0}'] = ''*18
    
    if num_anos == 2:
        header_r0 = header_r0[['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', f'Reparto % {y0}',
                               'ISS', 'PIS', 'IR']]
    elif num_anos == 3:
        header_r0 = header_r0[['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', f'Reparto % {y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                               'ISS', 'PIS', 'IR']]
    else:
        header_r0 = header_r0[['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '', f'Reparto % {y1}',              
                               f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                               f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', f'Reparto % {y0}',
                               f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                               f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                               f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                               'ISS', 'PIS', 'IR']]
        
    
    n = 9
    new_index = pd.RangeIndex(len(header_r0)*(n+1))
    header_r = pd.DataFrame(np.nan, index=new_index, columns=header_r0.columns)
    ids = np.arange(len(header_r0))*(n+1)
    header_r.loc[ids] = header_r0.values


    header_casc0 = tb_analitica[['MES','CENARIO']].drop_duplicates().sort_values(by=['MES']).reset_index(drop=True).T
    
    for (a,b,c) in zip(range(60), temp_ano, mes):
        header_casc0 = header_casc0.rename(columns={a: f'{b}{c}'})
    
    header_casc0['Cod'] = 'Valores em R$ Mil'
    header_casc0['Itens / Período'] = ''
    header_casc0[f'{y1}'] = header_casc0[f'{y1}12']
    header_casc0[f'{y0}'] = header_casc0[f'{y0}12']
    
    if num_anos == 3:
        header_casc0[f'{y01}'] = header_casc0[f'{y01}12']
    elif num_anos == 5:
        header_casc0[f'{y01}'] = header_casc0[f'{y01}12']
        header_casc0[f'{y02}'] = header_casc0[f'{y02}12']
        header_casc0[f'{y03}'] = header_casc0[f'{y03}12']
    
    if num_anos == 2:
        header_casc0 = header_casc0[['Cod','Itens / Período',
                                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}']]
    elif num_anos == 3:
        header_casc0 = header_casc0[['Cod','Itens / Período',
                                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}']]
    else:
        header_casc0 = header_casc0[['Cod','Itens / Período',
                                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}',
                                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}',
                                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}']]

    header_casc0 = header_casc0.iloc[1:]
    header_casc0.reset_index(drop = True, inplace=True)
    
    n = 9
    new_index = pd.RangeIndex(len(header_casc0)*(n+1))
    header_casc = pd.DataFrame(np.nan, index=new_index, columns=header_casc0.columns)
    ids = np.arange(len(header_casc0))*(n+1)
    header_casc.loc[ids] = header_casc0.values

    return header_p, header_r, header_s, header_casc

def ajustar_headers_finais(header_p, header_r, header_s, header_casc, pmodelo, pmodelop, rmodelo, smodelo, cmodelo, num_anos, cenario):
    '''Ajusta os headers finais em cima dos Dataframes de cada Modelo
    
    Parameters:
    header_p(Pandas DataFrame): Dataframe com o Header Final a ser utilizado para PModelos
    header_r(Pandas DataFrame): Dataframe com o Header Final a ser utilizado para RModelo
    header_s(Pandas DataFrame): Dataframe com o Header Final a ser utilizado para SModelo
    header_casc(Pandas DataFrame): Dataframe com o Header Final a ser utilizado para CascModelo
    pmodelo(Pandas DataFrame): PModelo final preenchido
    pmodelop(Pandas DataFrame): PModelo Ponta final preenchido
    rmodelo(Pandas DataFrame): RModelo final preenchido
    smodelo(Pandas DataFrame): SModelo final preenchido
    cmodelo(Pandas DataFrame): CascModelo final preenchido
    
    Returns:
    pmodelo_final(Pandas DataFrame): PModelo final preenchido com Header Final
    pmodelop_final(Pandas DataFrame): PModelo Ponta final preenchido com Header Final
    rmodelo_final(Pandas DataFrame): RModelo final preenchido com Header Final
    smodelo_final(Pandas DataFrame): SModelo final preenchido com Header Final
    cmodelo_final(Pandas DataFrame): CascModelo final preenchido com Header Final
    '''    
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    
    # Criar dataFrames vazios que receberão a concatenação do Modelo com o Header Final
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    pmodelo_ajust = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]
    
    pmodelop_ajust = deepcopy(pmodelo_ajust)
    rmodelo_ajust = deepcopy(pmodelo_ajust)
    smodelo_ajust = deepcopy(pmodelo_ajust)
    cmodelo_ajust = deepcopy(pmodelo_ajust)
    pmodelo_final0 = deepcopy(pmodelo_ajust)
    pmodelop_final0 = deepcopy(pmodelo_ajust)
    rmodelo_final0 = deepcopy(pmodelo_ajust)
    smodelo_final0 = deepcopy(pmodelo_ajust)
    cmodelo_final0 = deepcopy(pmodelo_ajust) 
    pmodelo_final = deepcopy(pmodelo_ajust)
    pmodelop_final = deepcopy(pmodelo_ajust)
    rmodelo_final = deepcopy(pmodelo_ajust)
    smodelo_final = deepcopy(pmodelo_ajust)
    cmodelo_final = deepcopy(pmodelo_ajust)   
    
    
    #Criar lista de meses e multiplicá-la por 5 (para 5 anos) e utilizar em loop for
    if num_anos == 2:
        mes = ['01','02','03','04','05','06','07','08','09','10','11','12','']
        temp_mes = list(mes)
        count = 2
        for i in range(count):
            for element in temp_mes:
                mes.append(element)
                
    elif num_anos == 3:
        mes = ['01','02','03','04','05','06','07','08','09','10','11','12','']
        temp_mes = list(mes)
        count = 3
        for i in range(count):
            for element in temp_mes:
                mes.append(element)
                
    else:
        mes = ['01','02','03','04','05','06','07','08','09','10','11','12','']
        temp_mes = list(mes)
        count = 5
        for i in range(count):
            for element in temp_mes:
                mes.append(element)

    #Criar lista de anos com 13 posições para cada ano e utilizar em loop for
    if num_anos == 2:
        ano = [y1, y0]
        temp_ano = []
        temp_ano += 13*[y1]
        temp_ano += 13*[y0]
        
        for i in range(len(pmodelo)):
            pmodelo_ajust[i] = pd.concat([header_p, pmodelo[i]])
            cols = pmodelo_ajust[i].columns
            header = pd.DataFrame(pmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                            8:'',
                                            9:' ',
                                           36:'ISS',
                                           37:'PIS',
                                           38:'IR'
                                           })
            for (a,b,c) in zip(range(10,36), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelo_final0[i] = pd.concat([pmodelo_ajust[i].iloc[:1], header, pmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelo_final0[i].rename(columns = pmodelo_final0[i].iloc[2],inplace=True)

            pmodelo_filt = pmodelo_final0[0].iloc[11:]
            pmodelo_val = pmodelo_filt.index
            pmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelo_id.append(j)

            pmodelo_final[i] = pmodelo_final0[i].reindex(pmodelo_id)


        for i in range(len(pmodelop)):
            pmodelop_ajust[i] = pd.concat([header_p, pmodelop[i]])
            cols = pmodelop_ajust[i].columns
            header = pd.DataFrame(pmodelop_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                            8:'',
                                            9:' ',
                                           36:'ISS',
                                           37:'PIS',
                                           38:'IR'
                                           })
            for (a,b,c) in zip(range(10,36), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelop_final0[i] = pd.concat([pmodelop_ajust[i].iloc[:1], header, pmodelop_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelop_final0[i].rename(columns = pmodelop_final0[i].iloc[2],inplace=True)

            pmodelop_filt = pmodelop_final0[0].iloc[11:]
            pmodelop_val = pmodelop_filt.index
            pmodelop_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelop_id.append(j)

            pmodelop_final[i] = pmodelop_final0[i].reindex(pmodelo_id)



        for i in range(len(rmodelo)):
            rmodelo_ajust[i] = pd.concat([header_r, rmodelo[i]])
            cols = rmodelo_ajust[i].columns
            header = pd.DataFrame(rmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Segmentos',
                                            8:'Linha excel',
                                            9:'',
                                           10:'Reparto % 2022',
                                           37:'Reparto % 2023',
                                           38:'ISS',
                                           39:'PIS',
                                           40:'IR'
                                           })
            for (a,b,c) in zip(range(11,37), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})
            rmodelo_final0[i] = pd.concat([rmodelo_ajust[i].iloc[:1], header, rmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            rmodelo_final0[i].rename(columns = rmodelo_final0[i].iloc[2],inplace=True)

            rmodelo_filt = rmodelo_final0[0].iloc[11:]
            rmodelo_val = rmodelo_filt.index
            rmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in rmodelo_val:
                rmodelo_id.append(j)

            rmodelo_final[i] = rmodelo_final0[i].reindex(rmodelo_id)        



        for i in range(len(smodelo)):
            smodelo_ajust[i] = pd.concat([header_s, smodelo[i]])
            cols = smodelo_ajust[i].columns
            header = pd.DataFrame(smodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                           34:'FL_DIAS_CORRIDOS',
                                           35:'ISS',
                                           36:'PIS',
                                           37:'IR'
                                           })
            for (a,b,c) in zip(range(8,34), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            smodelo_final0[i] = pd.concat([smodelo_ajust[i].iloc[:1], header, smodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            smodelo_final0[i].rename(columns = smodelo_final0[i].iloc[2],inplace=True)

            smodelo_filt = smodelo_final0[0].iloc[11:]
            smodelo_val = smodelo_filt.index
            smodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in smodelo_val:
                smodelo_id.append(j)

            smodelo_final[i] = smodelo_final0[i].reindex(smodelo_id)   


        for i in range(len(cmodelo)):
            cmodelo_ajust[i] = pd.concat([header_casc, cmodelo[i]])
            cols = cmodelo_ajust[i].columns
            header = pd.DataFrame(cmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Cod',
                                            1:'Itens / Período'
                                           })
            for (a,b,c) in zip(range(2,28), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            cmodelo_final0[i] = pd.concat([cmodelo_ajust[i].iloc[:1], header, cmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            cmodelo_final0[i].rename(columns = cmodelo_final0[i].iloc[2],inplace=True)

            cmodelo_filt = cmodelo_final0[0].iloc[11:]
            cmodelo_val = cmodelo_filt.index
            cmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in cmodelo_val:
                cmodelo_id.append(j)

            cmodelo_final[i] = cmodelo_final0[i].reindex(cmodelo_id)   
        
        
        
    elif num_anos == 3:
        ano = [y1, y0, y01]
        temp_ano = []
        temp_ano += 13*[y1]
        temp_ano += 13*[y0]
        temp_ano += 13*[y01]
        
        for i in range(len(pmodelo)):
            pmodelo_ajust[i] = pd.concat([header_p, pmodelo[i]])
            cols = pmodelo_ajust[i].columns
            header = pd.DataFrame(pmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                            8:'',
                                            9:' ',
                                           49:'ISS',
                                           50:'PIS',
                                           51:'IR'
                                           })
            for (a,b,c) in zip(range(10,49), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelo_final0[i] = pd.concat([pmodelo_ajust[i].iloc[:1], header, pmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelo_final0[i].rename(columns = pmodelo_final0[i].iloc[2],inplace=True)

            pmodelo_filt = pmodelo_final0[0].iloc[11:]
            pmodelo_val = pmodelo_filt.index
            pmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelo_id.append(j)

            pmodelo_final[i] = pmodelo_final0[i].reindex(pmodelo_id)


        for i in range(len(pmodelop)):
            pmodelop_ajust[i] = pd.concat([header_p, pmodelop[i]])
            cols = pmodelop_ajust[i].columns
            header = pd.DataFrame(pmodelop_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                            8:'',
                                            9:' ',
                                           49:'ISS',
                                           50:'PIS',
                                           51:'IR'
                                           })
            for (a,b,c) in zip(range(10,49), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelop_final0[i] = pd.concat([pmodelop_ajust[i].iloc[:1], header, pmodelop_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelop_final0[i].rename(columns = pmodelop_final0[i].iloc[2],inplace=True)

            pmodelop_filt = pmodelop_final0[0].iloc[11:]
            pmodelop_val = pmodelop_filt.index
            pmodelop_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelop_id.append(j)

            pmodelop_final[i] = pmodelop_final0[i].reindex(pmodelo_id)



        for i in range(len(rmodelo)):
            rmodelo_ajust[i] = pd.concat([header_r, rmodelo[i]])
            cols = rmodelo_ajust[i].columns
            header = pd.DataFrame(rmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Segmentos',
                                            8:'Linha excel',
                                            9:'',
                                           10:'Reparto % 2022',
                                           37:'Reparto % 2023',
                                           51:'ISS',
                                           52:'PIS',
                                           53:'IR'
                                           })
            for (a,b,c) in zip(range(11,37), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})
            for (a,b,c) in zip(range(38,54), temp_ano[26:], mes):
                header = header.rename(columns={a: f'{b}{c}'})
            rmodelo_final0[i] = pd.concat([rmodelo_ajust[i].iloc[:1], header, rmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            rmodelo_final0[i].rename(columns = rmodelo_final0[i].iloc[2],inplace=True)

            rmodelo_filt = rmodelo_final0[0].iloc[11:]
            rmodelo_val = rmodelo_filt.index
            rmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in rmodelo_val:
                rmodelo_id.append(j)

            rmodelo_final[i] = rmodelo_final0[i].reindex(rmodelo_id)        



        for i in range(len(smodelo)):
            smodelo_ajust[i] = pd.concat([header_s, smodelo[i]])
            cols = smodelo_ajust[i].columns
            header = pd.DataFrame(smodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                           47:'FL_DIAS_CORRIDOS',
                                           48:'ISS',
                                           49:'PIS',
                                           50:'IR'
                                           })
            for (a,b,c) in zip(range(8,47), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            smodelo_final0[i] = pd.concat([smodelo_ajust[i].iloc[:1], header, smodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            smodelo_final0[i].rename(columns = smodelo_final0[i].iloc[2],inplace=True)

            smodelo_filt = smodelo_final0[0].iloc[11:]
            smodelo_val = smodelo_filt.index
            smodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in smodelo_val:
                smodelo_id.append(j)

            smodelo_final[i] = smodelo_final0[i].reindex(smodelo_id)   


        for i in range(len(cmodelo)):
            cmodelo_ajust[i] = pd.concat([header_casc, cmodelo[i]])
            cols = cmodelo_ajust[i].columns
            header = pd.DataFrame(cmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Cod',
                                            1:'Itens / Período'
                                           })
            for (a,b,c) in zip(range(2,41), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            cmodelo_final0[i] = pd.concat([cmodelo_ajust[i].iloc[:1], header, cmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            cmodelo_final0[i].rename(columns = cmodelo_final0[i].iloc[2],inplace=True)

            cmodelo_filt = cmodelo_final0[0].iloc[11:]
            cmodelo_val = cmodelo_filt.index
            cmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in cmodelo_val:
                cmodelo_id.append(j)

            cmodelo_final[i] = cmodelo_final0[i].reindex(cmodelo_id)   
        
        
        
    else:
        ano = [y1, y0, y01, y02, y03]
        temp_ano = []
        temp_ano += 13*[y1]
        temp_ano += 13*[y0]
        temp_ano += 13*[y01]
        temp_ano += 13*[y02]
        temp_ano += 13*[y03]

        for i in range(len(pmodelo)):
            pmodelo_ajust[i] = pd.concat([header_p, pmodelo[i]])
            cols = pmodelo_ajust[i].columns
            header = pd.DataFrame(pmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Totalizador',
                                            1:'Chave Cascada',
                                            2:'Alíquotas',
                                            3:'Cascada',
                                            4:'Conceito',
                                            5:'Cod',
                                            6:'Itens / Período',
                                            7:'Novo Modelo',
                                            8:'',
                                            9:' ',
                                           75:'ISS',
                                           76:'PIS',
                                           77:'IR'
                                           })
            for (a,b,c) in zip(range(10,75), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelo_final0[i] = pd.concat([pmodelo_ajust[i].iloc[:1], header, pmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelo_final0[i].rename(columns = pmodelo_final0[i].iloc[2],inplace=True)

            pmodelo_filt = pmodelo_final0[0].iloc[11:]
            pmodelo_val = pmodelo_filt.index
            pmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelo_id.append(j)

            pmodelo_final[i] = pmodelo_final0[i].reindex(pmodelo_id)

        for i in range(len(pmodelop)):
            pmodelop_ajust[i] = pd.concat([header_p, pmodelop[i]])
            cols = pmodelop_ajust[i].columns
            header = pd.DataFrame(pmodelop_ajust[i].columns).T
            header = header.rename(columns={
                0:'Totalizador', 1:'Chave Cascada', 2:'Alíquotas', 3:'Cascada', 4:'Conceito', 5:'Cod', 6:'Itens / Período', 7:'Novo Modelo', 8:'', 9:' ', 75:'ISS', 76:'PIS', 77:'IR' 
            })
            for (a,b,c) in zip(range(10,75), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            pmodelop_final0[i] = pd.concat([pmodelop_ajust[i].iloc[:1], header, pmodelop_ajust[i].iloc[1:]]).reset_index(drop=True)
            pmodelop_final0[i].rename(columns = pmodelop_final0[i].iloc[2],inplace=True)

            pmodelop_filt = pmodelop_final0[0].iloc[11:]
            pmodelop_val = pmodelop_filt.index
            pmodelop_id = [2,3,4,5,6,7,8,9,0,1]
            for j in pmodelo_val:
                pmodelop_id.append(j)

            pmodelop_final[i] = pmodelop_final0[i].reindex(pmodelo_id)

        for i in range(len(rmodelo)):
            rmodelo_ajust[i] = pd.concat([header_r, rmodelo[i]])
            cols = rmodelo_ajust[i].columns
            header = pd.DataFrame(rmodelo_ajust[i].columns).T
            header = header.rename(columns={
                0:'Totalizador', 1:'Chave Cascada', 2:'Alíquotas', 3:'Cascada', 4:'Conceito', 5:'Cod', 6:'Itens / Período', 
                7:'Segmentos', 8:'Linha excel', 9:'', 10:'Reparto % 2022', 37:'Reparto % 2023', 77:'ISS', 78:'PIS', 79:'IR' 
            })
            for (a,b,c) in zip(range(11,37), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})
            for (a,b,c) in zip(range(38,77), temp_ano[26:], mes):
                header = header.rename(columns={a: f'{b}{c}'})
            rmodelo_final0[i] = pd.concat([rmodelo_ajust[i].iloc[:1], header, rmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            rmodelo_final0[i].rename(columns = rmodelo_final0[i].iloc[2],inplace=True)

            rmodelo_filt = rmodelo_final0[0].iloc[11:]
            rmodelo_val = rmodelo_filt.index
            rmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in rmodelo_val:
                rmodelo_id.append(j)

            rmodelo_final[i] = rmodelo_final0[i].reindex(rmodelo_id)        

        for i in range(len(smodelo)):
            smodelo_ajust[i] = pd.concat([header_s, smodelo[i]])
            cols = smodelo_ajust[i].columns
            header = pd.DataFrame(smodelo_ajust[i].columns).T
            header = header.rename(columns={
                0:'Totalizador', 1:'Chave Cascada', 2:'Alíquotas', 3:'Cascada', 4:'Conceito', 5:'Cod', 6:'Itens / Período', 7:'Novo Modelo', 73:'FL_DIAS_CORRIDOS', 74:'ISS', 75:'PIS', 76:'IR'
            })
            for (a,b,c) in zip(range(8,73), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            smodelo_final0[i] = pd.concat([smodelo_ajust[i].iloc[:1], header, smodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            smodelo_final0[i].rename(columns = smodelo_final0[i].iloc[2],inplace=True)

            smodelo_filt = smodelo_final0[0].iloc[11:]
            smodelo_val = smodelo_filt.index
            smodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in smodelo_val:
                smodelo_id.append(j)

            smodelo_final[i] = smodelo_final0[i].reindex(smodelo_id)   


        for i in range(len(cmodelo)):
            cmodelo_ajust[i] = pd.concat([header_casc, cmodelo[i]])
            cols = cmodelo_ajust[i].columns
            header = pd.DataFrame(cmodelo_ajust[i].columns).T
            header = header.rename(columns={0:'Cod', 1:'Itens / Período'})
            for (a,b,c) in zip(range(2,67), temp_ano, mes):
                header = header.rename(columns={a: f'{b}{c}'})

            cmodelo_final0[i] = pd.concat([cmodelo_ajust[i].iloc[:1], header, cmodelo_ajust[i].iloc[1:]]).reset_index(drop=True)
            cmodelo_final0[i].rename(columns = cmodelo_final0[i].iloc[2],inplace=True)

            cmodelo_filt = cmodelo_final0[0].iloc[11:]
            cmodelo_val = cmodelo_filt.index
            cmodelo_id = [2,3,4,5,6,7,8,9,0,1]
            for j in cmodelo_val:
                cmodelo_id.append(j)

            cmodelo_final[i] = cmodelo_final0[i].reindex(cmodelo_id)   
        
    return pmodelo_final, pmodelop_final, rmodelo_final, smodelo_final, cmodelo_final

def acrescentar_linhas_em_branco_renomear_colunas_data(pmodelo, pmodelop, rmodelo, smodelo, cascmodelo, num_anos, cenario):
    '''Acrescenta linhas em branco na estrutura dos Modelos antes de passar para Excel, bem como renomeia o nome das colunas de data.
    
    Parameters:
    pmodelo(Pandas DataFrame): PModelo final preenchido
    pmodelop(Pandas DataFrame): PModelo Ponta final preenchido
    rmodelo(Pandas DataFrame): RModelo final preenchido
    smodelo(Pandas DataFrame): SModelo final preenchido
    cascmodelo(Pandas DataFrame): CascModelo final preenchido
    
    Returns:
    pmodelo(Pandas DataFrame): PModelo final ajustado
    pmodelop(Pandas DataFrame): PModelo Ponta final ajustado
    rmodelo(Pandas DataFrame): RModelo final ajustado
    smodelo(Pandas DataFrame): SModelo final ajustado
    cascmodelo(Pandas DataFrame): CascModelo final ajustado
    '''  
    # Get ano atual (y0), ano anterior (y1), ano seguinte (y01), ano seguinte posterior (y02), ano seguinte posterior posterior (y03)
    m0 = datetime.now().month
    # if m0 == 1 and cenario.lower() in ('avance', 'avanc', 'avan', 'fechto', 'fechamento', 'fecham', 'p27'):
    #     y0 = datetime.now().year - 1
    # else:
    #     y0 = datetime.now().year
    y0 = datetime.now().year - 1
    y1 = y0 - 1
    y01 = y0 + 1
    y02 = y0 + 2
    y03 = y0 + 3
    y0 = str(y0)
    y1 = str(y1)
    y01 = str(y01)
    y02 = str(y02)
    y03 = str(y03)
    
    pmodelo_temp = deepcopy(pmodelo)
    rmodelo_temp = deepcopy(rmodelo)
    pmodelop_temp = deepcopy(pmodelop)
    smodelo_temp = deepcopy(smodelo)
    cascmodelo_temp = deepcopy(cascmodelo)
    
    # Criar colunas em branco (estética)
    for i in range(len(pmodelo)):
        pmodelo_temp[i]['-'] = '-' 
        pmodelop_temp[i]['-'] = '-'        
        rmodelo_temp[i]['-'] = '-'         
        smodelo_temp[i]['-'] = '-'         
        cascmodelo_temp[i]['-'] = '-'

    # Criar um header temporário
    for i in range(len(pmodelo)):
        pmodelo_temp[i].columns = pmodelo_temp[i].iloc[9]       
        pmodelop_temp[i].columns = pmodelop_temp[i].iloc[9]  
        rmodelo_temp[i].columns = rmodelo_temp[i].iloc[9]
        smodelo_temp[i].columns = smodelo_temp[i].iloc[9]
        cascmodelo_temp[i].columns = cascmodelo_temp[i].iloc[9]
        
    # Seleciona quais colunas serão utilizadas e em qual ordem de cada Modelo
    if num_anos == 2:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '',' ','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     '-','-','-','-','-','-','-','ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '','-', f'Reparto % {y1}', 
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-', f'Reparto % {y0}',
                     '-','-','-','-', 'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}',
                     '-','-','FL_DIAS_CORRIDOS','-','-','-','-','ISS','PIS','IR']

        ccols_val = ['-','-','-','-','-','Cod','Itens / Período','-','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}']
        
    elif num_anos == 3:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '',' ','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     '-','-','-','-','-','-','-','ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '','-', f'Reparto % {y1}', 
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', '-',
                     f'Reparto % {y0}','-','-','-','-', 'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}',
                     '-','-','FL_DIAS_CORRIDOS','-','-','-','-','ISS','PIS','IR']

        ccols_val = ['-','-','-','-','-','Cod','Itens / Período','-','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}']
        
    else:
        pcols_val = ['Totalizador','Chave Cascada','Alíquotas','Cascada','Conceito','Cod','Itens / Período','Novo Modelo', '',' ','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', '-',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}', '-',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                     '-','-','-','-','-','-','-','ISS', 'PIS', 'IR']

        rcols_val = ['Totalizador', 'Chave Cascada', 'Alíquotas', 'Cascada', 'Conceito', 'Cod', 'Itens / Período', 'Segmentos','Linha excel', '','-', f'Reparto % {y1}', 
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', '-',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}', '-',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}', '-',
                     f'Reparto % {y0}','-','-','-','-', 'ISS', 'PIS', 'IR']

        scols_val = ['Totalizador','Chave Cascada','Alíquotas','Conceito','Cascada','Cod','Itens / Período', 'Novo Modelo','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', '-',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}', '-',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}',
                     '-','-','FL_DIAS_CORRIDOS','-','-','-','-','ISS','PIS','IR']

        ccols_val = ['-','-','-','-','-','Cod','Itens / Período','-','-','-','-','-',
                     f'{y1}01',f'{y1}02',f'{y1}03',f'{y1}04',f'{y1}05',f'{y1}06',f'{y1}07',f'{y1}08',f'{y1}09',f'{y1}10',f'{y1}11',f'{y1}12',f'{y1}', '-',
                     f'{y0}01',f'{y0}02',f'{y0}03',f'{y0}04',f'{y0}05',f'{y0}06',f'{y0}07',f'{y0}08',f'{y0}09',f'{y0}10',f'{y0}11',f'{y0}12',f'{y0}', '-',
                     f'{y01}01',f'{y01}02',f'{y01}03',f'{y01}04',f'{y01}05',f'{y01}06',f'{y01}07',f'{y01}08',f'{y01}09',f'{y01}10',f'{y01}11',f'{y01}12',f'{y01}', '-',
                     f'{y02}01',f'{y02}02',f'{y02}03',f'{y02}04',f'{y02}05',f'{y02}06',f'{y02}07',f'{y02}08',f'{y02}09',f'{y02}10',f'{y02}11',f'{y02}12',f'{y02}', '-',
                     f'{y03}01',f'{y03}02',f'{y03}03',f'{y03}04',f'{y03}05',f'{y03}06',f'{y03}07',f'{y03}08',f'{y03}09',f'{y03}10',f'{y03}11',f'{y03}12',f'{y03}']
    
    df1=df2=df3=df4=df5=df6=df7=df8=df9=df10=df11=df12=df13=df14=df15=df16=df17=df18=df19=pd.DataFrame()
    pmodelo_final = [df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,df17,df18,df19]
    rmodelo_final = deepcopy(pmodelo_final)
    pmodelop_final = deepcopy(pmodelo_final)
    smodelo_final = deepcopy(pmodelo_final)
    cascmodelo_final = deepcopy(pmodelo_final)
    
    # Filtra os dataframes pelas colunas assinaladas acima
    for i in range(len(pmodelo)):
        pmodelo_final[i] = pmodelo_temp[i][pcols_val]
        pmodelop_final[i] = pmodelop_temp[i][pcols_val]
        rmodelo_final[i] = rmodelo_temp[i][rcols_val]
        smodelo_final[i] = smodelo_temp[i][scols_val]
        cascmodelo_final[i] = cascmodelo_temp[i][ccols_val]
        
    # Voltar com o header final
    for i in range(len(pmodelo)):
        pmodelo_final[i] = pmodelo_final[i].replace('-', '') 
        pmodelop_final[i] = pmodelop_final[i].replace('-', '')
        rmodelo_final[i] = rmodelo_final[i].replace('-', '')
        smodelo_final[i] = smodelo_final[i].replace('-', '')
        cascmodelo_final[i] = cascmodelo_final[i].replace('-', '')
        
    for i in range(len(pmodelo)):
        pmodelo_final[i].columns = pmodelo_final[i].iloc[2]       
        pmodelop_final[i].columns = pmodelop_final[i].iloc[2]  
        rmodelo_final[i].columns = rmodelo_final[i].iloc[2]
        smodelo_final[i].columns = smodelo_final[i].iloc[2]
        cascmodelo_final[i].columns = cascmodelo_final[i].iloc[2]
    
    return pmodelo_final, pmodelop_final, rmodelo_final, smodelo_final, cascmodelo_final

def gerar_excel_var_atual(df_pnl_Var, df_pnl_VarTotal, pmodelo, pmodelop, rmodelo, smodelo, cascmodelo, segmentos, cenario, mes_ano_ref, versao):
    '''Gera todos os Modelos em Excel, em formato raw.
    
    Parameters:
    pmodelo_final(Pandas DataFrame): PModelo final preenchido
    pmodelop_final(Pandas DataFrame): PModelo Ponta final preenchido
    smodelo_final(Pandas DataFrame): SModelo final preenchido
    cascmodelo_final(Pandas DataFrame): CascModelo final preenchido
    segmentos(dict): Segmentos oriundos do arquivo de variáveis (variaveis_var_atual)
    cenario(str): Cenário utilizado (Prévia, Avance, Fchto) oriundo do arquivo de variáveis (variaveis_var_atual)
    mes_ano_ref(str): Mes e ano de referência, oriundo do arquivo de variáveis (variaveis_var_atual)
    versao(str): Versão de referência, oriundo do arquivo de variáveis (variaveis_var_atual)
    
    '''       
    path_excel_raw = f"Novo_Varejo_Atual_{cenario}_{mes_ano_ref}_v{versao}_raw.xlsx"
    with pd.ExcelWriter(path_excel_raw, engine="xlsxwriter") as writer:
        # BAI_check.to_excel(writer, sheet_name = 'valid_BAI_RvsCasc', index = False)
        # Segto_check.to_excel(writer, sheet_name = 'valid_TotSegto_R', index = False)
        # Segto_P_check.to_excel(writer, sheet_name = 'valid_TotSegto_P', index = False)
        df_pnl_VarTotal.to_excel(writer, sheet_name = 'PNL_VarTotal', index = False)
        df_pnl_Var.to_excel(writer, sheet_name = 'PNL_Var', index = False)
        
        
        cont1 = 0
        for i in smodelo:
            i.to_excel(writer, sheet_name = f'S_{list(segmentos.values())[cont4]}', index = False)
            cont4 += 1
            
        cont2 = 0
        for i in pmodelop:
            i.to_excel(writer, sheet_name = f'P_{list(segmentos.values())[cont2]}_Ponta', index = False)
            cont2 += 1    
            
        cont3 = 0
        for i in pmodelo:
            i.to_excel(writer, sheet_name = f'P_{list(segmentos.values())[cont1]}', index = False)
            cont1 += 1
        
        cont4 = 0
        for i in rmodelo:
            i.to_excel(writer, sheet_name = f'R_{list(segmentos.values())[cont3]}', index = False)
            cont3 += 1   
        
        cont5 = 0
        for i in cascmodelo:
            i.to_excel(writer, sheet_name = f'Casc_{list(segmentos.values())[cont5]}', index = False)
            cont5 += 1    
    return path_excel_raw

def personalizar_excel(path_excel_raw, cenario, mes_ano_ref, versao, num_anos):
    '''Personalizar o arquivo excel raw esteticamente de acordo com o Modelo BP.
    Parameters:
    path_excel_raw(str): Path do arquivo Excel raw
    cenario(str): Cenário utilizado (Prévia, Avance, Fchto) oriundo do arquivo de variáveis (variaveis_var_atual)
    mes_ano_ref(str): Mes e ano de referência, oriundo do arquivo de variáveis (variaveis_var_atual)
    versao(str): Versão de referência, oriundo do arquivo de variáveis (variaveis_var_atual)
    '''  
    path_excel = f"Novo_Varejo_Atual_{cenario}_{mes_ano_ref}_v{versao}.xlsx"
    
    cor_tabs = {
        'pmodelo': '808080', 'pmodelop': '000000', 'rmodelo': 'FF0000', 'cascmodelo': '305496'  
    }
    p_cod_cinza_medio_escuro = [78, 91, 100, 191, 193, 198, 204, 206, 207, 2 , 215, 234, 302, 307, 311]
    r_cod_cinza_medio_escuro = [
        78, 91, 100, 191, 193, 198, 204, 206, 207, 211, 215, 234, 302, 307, 311, 313, 334, 352, 373, 383, 395, 403, 432, 493, 503,
        575, 583, 600, 613, 623, 630, 643, 661, 667, 677, 690, 705, 731, 737, 757, 778, 816, 828, 848, 858, 870, 884, 895, 910,
        920, 942, 966, 990, 1008, 1020, 1064, 1081, 1109, 1119, 1127, 1150, 1161, 1171, 1178, 1263, 1264, 1265, 1266, 701098
    ]
    r_cod_cinza_medio = [1216, 1217, 1218, 1219, 1220, 1221]
    r_cod_cinza_escuro = [1262]
    r_cod_vermelho = [579, 581, 880, 882, 1157, 1159, 1198, 1204]
    s_cod_cinza_medio_escuro = [78, 91, 100, 191, 193, 198, 204, 206, 207, 211, 215, 234, 302, 307, 311]
    c_cod_cinza_medio_escuro = [122, 123, 124, 125, 126, 127]
    c_cod_cinza_medio = [131, 132, 133, 134]
    c_cod_cinza_escuro = [130]
    c_cod_vermelho = [66, 67, 92, 93, 111, 115, 118]
    null_value = [-999]
    
    backgroud_vermelho = PatternFill(patternType='solid', fgColor='FF0000')
    
    backgroud_cinza_medio = PatternFill(patternType='solid', fgColor='BFBFBF')
    
    backgroud_cinza_medio_escuro = PatternFill(patternType='solid', fgColor='C0C0C0')
    
    backgroud_cinza_escuro = PatternFill(patternType='solid', fgColor='595959')
    
    fonte_branca_negrito = Font(name='Calibri', size=11, bold=True, italic=False, strike=False, underline='none', color='FFFFFF')

    fonte_preta_negrito = Font(name='Calibri', size=11, bold=True, italic=False, strike=False, underline='none', color='000000')

    wb = load_workbook(path_excel_raw)
    sheet_names = wb.sheetnames
    
    if num_anos == 2:
        p_head = [
            'A11', 'B11', 'C11', 'D11', 'E11', 'F11', 'G11', 'M11', 'N11', 'O11', 'P11', 'Q11', 'R11', 'S11', 'T11', 'U11', 'V11', 'W11', 'X11', 
            'Y11', 'AA11', 'AB11', 'AC11', 'AD11', 'AE11', 'AF11', 'AG11', 'AH11', 'AI11', 'AJ11', 'AK11', 'AL11', 'AM11', 'AU11', 'AV11', 'AW11'
        ]

        r_head = ['A11','B11','C11','D11','E11','F11','G11','H11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11',
                  'AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11',
                  'AO11','AT11','AU11','AV11']

        s_head = ['A11','B11','C11','D11','E11','F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11',
                  'AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11',
                  'AP11','AU11','AV11','AW11']

        c_head = ['F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11', 'Y11',
                  'AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11']

        p_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X','Y',
                      'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']

        r_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X','Y',
                      'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']

        p_cols_perc = ['AU','AV','AW']

        r_cols_perc = ['L','AO','AU','AV','AW']

        s_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X','Y',
                      'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']

        c_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X','Y',
                      'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']
        
        for i in sheet_names:
            ws = wb[i]

            if '_Ponta' in i:
                
                # Converter colunas numéricas para Contábil
                for a in p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')
                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelop')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell   
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70

                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 40
                while column < 47:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 47
                while column < 50:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                for col in ['H', 'I', 'J', 'L', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros

                if num_anos == 2:
                    tamanho_tb = 'A11:AW3670'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK3670'
                else:
                    tamanho_tb = 'A11:CM3670'
                ws.auto_filter.ref = tamanho_tb

            elif 'R_' in i:


                # Converter colunas numéricas para Contábil
                for a in  r_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  r_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('rmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in r_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in r_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in r_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['I11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['I11'].fill = backgroud_cinza_medio_escuro
                ws['I11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 42
                while column < 46:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1   

                column = 46
                while column < 49:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1  

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 7 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 11.7 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[1360].height = 4.5 + 0.7
                ws.row_dimensions[1482].height = 4.5 + 0.7
                ws.row_dimensions[1491].height = 4.5 + 0.7
                ws.row_dimensions[1493].height = 4.5 + 0.7
                ws.row_dimensions[2181].height = 4.5 + 0.7
                ws.row_dimensions[2183].height = 4.5 + 0.7
                ws.row_dimensions[2185].height = 4.5 + 0.7
                ws.row_dimensions[2752].height = 4.5 + 0.7
                ws.row_dimensions[2754].height = 4.5 + 0.7
                ws.row_dimensions[2756].height = 4.5 + 0.7
                ws.row_dimensions[2819].height = 4.5 + 0.7
                ws.row_dimensions[2821].height = 4.5 + 0.7
                ws.row_dimensions[2830].height = 4.5 + 0.7
                ws.row_dimensions[2905].height = 4.5 + 0.7
                ws.row_dimensions[2906].height = 4.5 + 0.7
                ws.row_dimensions[2908].height = 4.5 + 0.7
                ws.row_dimensions[4036].height = 4.5 + 0.7
                ws.row_dimensions[4270].height = 4.5 + 0.7
                ws.row_dimensions[4303].height = 4.5 + 0.7
                ws.row_dimensions[4613].height = 4.5 + 0.7
                ws.row_dimensions[5001].height = 4.5 + 0.7
                ws.row_dimensions[5123].height = 4.5 + 0.7
                ws.row_dimensions[5132].height = 4.5 + 0.7
                ws.row_dimensions[5134].height = 4.5 + 0.7
                ws.row_dimensions[5136].height = 4.5 + 0.7
                ws.row_dimensions[5822].height = 4.5 + 0.7
                ws.row_dimensions[5824].height = 4.5 + 0.7
                ws.row_dimensions[5826].height = 4.5 + 0.7
                ws.row_dimensions[6393].height = 4.5 + 0.7
                ws.row_dimensions[6395].height = 4.5 + 0.7
                ws.row_dimensions[6397].height = 4.5 + 0.7
                ws.row_dimensions[6460].height = 4.5 + 0.7
                ws.row_dimensions[6462].height = 4.5 + 0.7
                ws.row_dimensions[6471].height = 4.5 + 0.7
                ws.row_dimensions[6546].height = 4.5 + 0.7
                ws.row_dimensions[6547].height = 4.5 + 0.7
                ws.row_dimensions[6549].height = 4.5 + 0.7
                ws.row_dimensions[7707].height = 4.5 + 0.7
                ws.row_dimensions[7738].height = 4.5 + 0.7
                ws.row_dimensions[7740].height = 4.5 + 0.7
                ws.row_dimensions[7774].height = 4.5 + 0.7
                ws.row_dimensions[8050].height = 4.5 + 0.7
                ws.row_dimensions[8438].height = 4.5 + 0.7
                ws.row_dimensions[8560].height = 4.5 + 0.7
                ws.row_dimensions[8569].height = 4.5 + 0.7
                ws.row_dimensions[8571].height = 4.5 + 0.7
                ws.row_dimensions[8573].height = 4.5 + 0.7
                ws.row_dimensions[9259].height = 4.5 + 0.7
                ws.row_dimensions[9261].height = 4.5 + 0.7
                ws.row_dimensions[9263].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='L', end='X', hidden=False) 
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                for col in ['H', 'I', 'J', 'L', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:AV9937'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK9937'
                else:
                    tamanho_tb = 'A11:CM9937'
                ws.auto_filter.ref = tamanho_tb

            elif 'S_' in i:            

                # Converter colunas numéricas para porcentagem
                for a in s_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')        


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in s_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in s_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 40
                while column < 42:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 5 + 0.7
                    column += 1   

                column = 43
                while column < 47:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1   

                column = 47
                while column < 50:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1   

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AP'].width = 15 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                for col in ['H', 'I', 'J', 'L']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:AW972'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK972'
                else:
                    tamanho_tb = 'A11:CM972'
                ws.auto_filter.ref = tamanho_tb

            elif 'Casc_' in i:

                # Converter colunas numéricas para Contábil
                for a in  c_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('cascmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = 500, min_col = 1, max_col = 81):
                    cell_value = row[5].value
                    if cell_value in c_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in c_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in c_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1


                # set the width of the column
                ws.column_dimensions['A'].width = 0.7
                ws.column_dimensions['B'].width = 0.7
                ws.column_dimensions['C'].width = 0.7
                ws.column_dimensions['D'].width = 0.7
                ws.column_dimensions['E'].width = 0.7
                ws.column_dimensions['F'].width = 16 + 0.7
                ws.column_dimensions['G'].width = 50 + 0.7
                ws.column_dimensions['H'].width = 1.86 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[39].height = 4.5 + 0.7
                ws.row_dimensions[66].height = 4.5 + 0.7
                ws.row_dimensions[69].height = 4.5 + 0.7
                ws.row_dimensions[92].height = 4.5 + 0.7
                ws.row_dimensions[95].height = 4.5 + 0.7
                ws.row_dimensions[111].height = 4.5 + 0.7
                ws.row_dimensions[113].height = 4.5 + 0.7
                ws.row_dimensions[118].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                for col in ['A', 'B', 'C', 'D']:
                    ws.column_dimensions[col].hidden= True

               #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:AM135'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BA135'
                else:
                    tamanho_tb = 'A11:CC135'
                ws.auto_filter.ref = tamanho_tb

            elif 'PNL_' in i: 
                pass
            
            elif 'anl' in i:
                ws.auto_filter.ref = ws.dimensions

            else:
                #Tab PModelo           

                # Converter colunas numéricas para Contábil
                for a in  p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(3, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 40
                while column < 47:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 47
                while column < 50:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                for col in ['H', 'I', 'J', 'L', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:AW3667'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK3667'
                else:
                    tamanho_tb = 'A11:CM3667'
                ws.auto_filter.ref = tamanho_tb

        wb.save(path_excel)
        
    elif num_anos == 3:
        p_head = ['A11','B11','C11','D11','E11','F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                  'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AO11','AP11','AQ11',
                  'AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BI11','BJ11','BK11']

        r_head = ['A11','B11','C11','D11','E11','F11','G11','H11','L11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                  'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AO11','AP11','AQ11',
                  'AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BC11','BH11','BI11','BJ11']

        s_head = ['A11','B11','C11','D11','E11','F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                      'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AO11','AP11','AQ11','AR11','AS11',
                      'AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BD11','BI11','BJ11','BK11']

        c_head = ['F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                  'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11',
                 'AO11','AP11','AQ11','AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11']

        p_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AO','AP','AQ',
                      'AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA']

        r_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AO','AP','AQ',
                      'AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA']

        p_cols_perc = ['BI','BJ','BK']

        r_cols_perc = ['L','BC','BH','BI','BJ']

        s_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AO','AP','AQ','AR','AS',
                      'AT','AU','AV','AW','AX','AY','AZ','BA','BI','BJ','BK']

        c_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM',
                      'AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA']
        
        
        for i in sheet_names:
            ws = wb[i]

            if '_Ponta' in i:

                # Converter colunas numéricas para Contábil
                for a in p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')
                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelop')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell   
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 55
                while column < 61:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 61
                while column < 64:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 10 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'BI', 'BJ', 'BK']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                tamanho_tb = 'A11:BK3670'
                ws.auto_filter.ref = tamanho_tb

            elif 'R_' in i:


                # Converter colunas numéricas para Contábil
                for a in  r_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  r_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('rmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in r_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in r_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in r_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['I11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['I11'].fill = backgroud_cinza_medio_escuro
                ws['I11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 56
                while column < 60:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1  

                column = 60
                while column < 63:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 10 + 0.7
                    column += 1  

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 7 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 11.7 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 7 + 0.7
                ws.column_dimensions['BB'].width = 7 + 0.7
                ws.column_dimensions['BC'].width = 15 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[1360].height = 4.5 + 0.7
                ws.row_dimensions[1482].height = 4.5 + 0.7
                ws.row_dimensions[1491].height = 4.5 + 0.7
                ws.row_dimensions[1493].height = 4.5 + 0.7
                ws.row_dimensions[2181].height = 4.5 + 0.7
                ws.row_dimensions[2183].height = 4.5 + 0.7
                ws.row_dimensions[2185].height = 4.5 + 0.7
                ws.row_dimensions[2752].height = 4.5 + 0.7
                ws.row_dimensions[2754].height = 4.5 + 0.7
                ws.row_dimensions[2756].height = 4.5 + 0.7
                ws.row_dimensions[2819].height = 4.5 + 0.7
                ws.row_dimensions[2821].height = 4.5 + 0.7
                ws.row_dimensions[2830].height = 4.5 + 0.7
                ws.row_dimensions[2905].height = 4.5 + 0.7
                ws.row_dimensions[2906].height = 4.5 + 0.7
                ws.row_dimensions[2908].height = 4.5 + 0.7
                ws.row_dimensions[4036].height = 4.5 + 0.7
                ws.row_dimensions[4270].height = 4.5 + 0.7
                ws.row_dimensions[4303].height = 4.5 + 0.7
                ws.row_dimensions[4613].height = 4.5 + 0.7
                ws.row_dimensions[5001].height = 4.5 + 0.7
                ws.row_dimensions[5123].height = 4.5 + 0.7
                ws.row_dimensions[5132].height = 4.5 + 0.7
                ws.row_dimensions[5134].height = 4.5 + 0.7
                ws.row_dimensions[5136].height = 4.5 + 0.7
                ws.row_dimensions[5822].height = 4.5 + 0.7
                ws.row_dimensions[5824].height = 4.5 + 0.7
                ws.row_dimensions[5826].height = 4.5 + 0.7
                ws.row_dimensions[6393].height = 4.5 + 0.7
                ws.row_dimensions[6395].height = 4.5 + 0.7
                ws.row_dimensions[6397].height = 4.5 + 0.7
                ws.row_dimensions[6460].height = 4.5 + 0.7
                ws.row_dimensions[6462].height = 4.5 + 0.7
                ws.row_dimensions[6471].height = 4.5 + 0.7
                ws.row_dimensions[6546].height = 4.5 + 0.7
                ws.row_dimensions[6547].height = 4.5 + 0.7
                ws.row_dimensions[6549].height = 4.5 + 0.7
                ws.row_dimensions[7707].height = 4.5 + 0.7
                ws.row_dimensions[7738].height = 4.5 + 0.7
                ws.row_dimensions[7740].height = 4.5 + 0.7
                ws.row_dimensions[7774].height = 4.5 + 0.7
                ws.row_dimensions[8050].height = 4.5 + 0.7
                ws.row_dimensions[8438].height = 4.5 + 0.7
                ws.row_dimensions[8560].height = 4.5 + 0.7
                ws.row_dimensions[8569].height = 4.5 + 0.7
                ws.row_dimensions[8571].height = 4.5 + 0.7
                ws.row_dimensions[8573].height = 4.5 + 0.7
                ws.row_dimensions[9259].height = 4.5 + 0.7
                ws.row_dimensions[9261].height = 4.5 + 0.7
                ws.row_dimensions[9263].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='L', end='X', hidden=False) 
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'BH','BI', 'BJ']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                tamanho_tb = 'A11:BJ9937'
                ws.auto_filter.ref = tamanho_tb

            elif 'S_' in i:            

                # Converter colunas numéricas para porcentagem
                for a in s_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')        


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in s_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in s_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell                
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 54
                while column < 56:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1   

                column = 57
                while column < 61:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1   

                column = 61
                while column < 64:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 10 + 0.7
                    column += 1   

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BD'].width = 15 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                for col in ['H', 'I', 'J', 'L']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                tamanho_tb = 'A11:BK972'
                ws.auto_filter.ref = tamanho_tb

            elif 'Casc_' in i:

                # Converter colunas numéricas para Contábil
                for a in  c_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('cascmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = 500, min_col = 1, max_col = 81):
                    cell_value = row[5].value
                    if cell_value in c_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in c_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in c_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 0.7
                ws.column_dimensions['B'].width = 0.7
                ws.column_dimensions['C'].width = 0.7
                ws.column_dimensions['D'].width = 0.7
                ws.column_dimensions['E'].width = 0.7
                ws.column_dimensions['F'].width = 16 + 0.7
                ws.column_dimensions['G'].width = 50 + 0.7
                ws.column_dimensions['H'].width = 1.86 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 3 + 0.7
                ws.column_dimensions['AN'].width = 3 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[39].height = 4.5 + 0.7
                ws.row_dimensions[66].height = 4.5 + 0.7
                ws.row_dimensions[69].height = 4.5 + 0.7
                ws.row_dimensions[92].height = 4.5 + 0.7
                ws.row_dimensions[95].height = 4.5 + 0.7
                ws.row_dimensions[111].height = 4.5 + 0.7
                ws.row_dimensions[113].height = 4.5 + 0.7
                ws.row_dimensions[118].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                for col in ['A', 'B', 'C', 'D']:
                    ws.column_dimensions[col].hidden= True

               #Inserir filtros
                tamanho_tb = 'A11:BA135'
                ws.auto_filter.ref = tamanho_tb

            elif 'PNL_' in i: 
                pass
            
            elif 'anl' in i:
                ws.auto_filter.ref = ws.dimensions

            else:
                #Tab PModelo           

                # Converter colunas numéricas para Contábil
                for a in  p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(3, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 55
                while column < 61:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 61
                while column < 64:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 10 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'BI', 'BJ', 'BK']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                tamanho_tb = 'A11:BK3667'
                ws.auto_filter.ref = tamanho_tb

        wb.save(path_excel)
        
    else:
        p_head = ['A11','B11','C11','D11','E11','F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                  'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AO11','AP11','AQ11',
                  'AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BC11','BD11','BE11','BF11','BG11','BH11','BI11',
                  'BJ11','BK11','BL11','BM11','BN11','BO11','BQ11','BR11','BS11','BT11','BU11','BV11','BW11','BX11','BY11','BZ11','CA11',
                  'CB11','CC11','CK11','CL11','CM11']

        r_head = ['A11','B11','C11','D11','E11','F11','G11','H11','L11',
                  'M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11','Y11',
                  'AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11',
                  'AO11','AP11','AQ11','AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11', 
                  'BC11','BD11','BE11','BF11','BG11','BH11','BI11','BJ11','BK11','BL11','BM11','BN11','BO11', 
                  'BQ11','BR11','BS11','BT11','BU11','BV11','BW11','BX11','BY11','BZ11','CA11','CB11','CC11',
                  'CE11','CJ11','CK11','CL11']

        s_head = ['A11','B11','C11','D11','E11','F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                      'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11','AO11','AP11','AQ11','AR11','AS11',
                      'AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BC11','BD11','BE11','BF11','BG11','BH11','BI11',
                      'BJ11','BK11','BL11','BM11','BN11','BO11','BQ11','BR11','BS11','BT11','BU11','BV11','BW11','BX11','BY11','BZ11','CA11',
                      'CB11','CC11','CF11','CK11','CL11','CM11']

        c_head = ['F11','G11','M11','N11','O11','P11','Q11','R11','S11','T11','U11','V11','W11','X11',
                  'Y11','AA11','AB11','AC11','AD11','AE11','AF11','AG11','AH11','AI11','AJ11','AK11','AL11','AM11',
                 'AO11','AP11','AQ11','AR11','AS11','AT11','AU11','AV11','AW11','AX11','AY11','AZ11','BA11','BC11',
                  'BD11','BE11','BF11','BG11','BH11','BI11','BJ11','BK11','BL11','BM11','BN11','BO11','BQ11','BR11',
                  'BS11','BT11','BU11','BV11','BW11','BX11','BY11','BZ11','CA11','CB11','CC11']

        p_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AO','AP','AQ',
                      'AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BC','BD','BE','BF','BG','BH','BI',
                      'BJ','BK','BL','BM','BN','BO','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA',
                      'CB','CC']

        r_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X','Y',
                      'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM',
                      'AO', 'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA',
                      'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO',
                      'BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC',
                      'CE']

        p_cols_perc = ['CK','CL','CM']

        r_cols_perc = ['L','CE','CJ','CK','CL']

        s_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AO','AP','AQ','AR','AS',
                      'AT','AU','AV','AW','AX','AY','AZ','BA','BC','BD','BE','BF','BG','BH','BI',
                      'BJ','BK','BL','BM','BN','BO','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA',
                      'CB','CC','CK','CL','CM']

        c_cols_num = ['M','N','O','P','Q','R','S','T','U','V','W','X',
                      'Y','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM',
                      'AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BC',
                      'BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BQ','BR',
                      'BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC']
    

        for i in sheet_names:
            ws = wb[i]

            if '_Ponta' in i:

                # Converter colunas numéricas para Contábil
                for a in p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')
                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelop')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell   
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 55
                while column < 68:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 69
                while column < 82:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 82
                while column < 89:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 89
                while column < 92:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7
                ws.column_dimensions['BP'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                ws.column_dimensions.group(start='BC', end='BN', hidden=True)
                ws.column_dimensions.group(start='BQ', end='CB', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros

                if num_anos == 2:
                    tamanho_tb = 'A11:CM3670'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK3670'
                else:
                    tamanho_tb = 'A11:CM3670'
                ws.auto_filter.ref = tamanho_tb

            elif 'R_' in i:


                # Converter colunas numéricas para Contábil
                for a in  r_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  r_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('rmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in r_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in r_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in r_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in r_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['I11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['I11'].fill = backgroud_cinza_medio_escuro
                ws['I11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['H12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 55
                while column < 68:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 69
                while column < 82:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1  

                column = 84
                while column < 89:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1  

                column = 88
                while column < 91:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1  

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 7 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 11.7 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7
                ws.column_dimensions['BP'].width = 5 + 0.7
                ws.column_dimensions['CD'].width = 5 + 0.7
                ws.column_dimensions['CE'].width = 15 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[1360].height = 4.5 + 0.7
                ws.row_dimensions[1482].height = 4.5 + 0.7
                ws.row_dimensions[1491].height = 4.5 + 0.7
                ws.row_dimensions[1493].height = 4.5 + 0.7
                ws.row_dimensions[2181].height = 4.5 + 0.7
                ws.row_dimensions[2183].height = 4.5 + 0.7
                ws.row_dimensions[2185].height = 4.5 + 0.7
                ws.row_dimensions[2752].height = 4.5 + 0.7
                ws.row_dimensions[2754].height = 4.5 + 0.7
                ws.row_dimensions[2756].height = 4.5 + 0.7
                ws.row_dimensions[2819].height = 4.5 + 0.7
                ws.row_dimensions[2821].height = 4.5 + 0.7
                ws.row_dimensions[2830].height = 4.5 + 0.7
                ws.row_dimensions[2905].height = 4.5 + 0.7
                ws.row_dimensions[2906].height = 4.5 + 0.7
                ws.row_dimensions[2908].height = 4.5 + 0.7
                ws.row_dimensions[4036].height = 4.5 + 0.7
                ws.row_dimensions[4270].height = 4.5 + 0.7
                ws.row_dimensions[4303].height = 4.5 + 0.7
                ws.row_dimensions[4613].height = 4.5 + 0.7
                ws.row_dimensions[5001].height = 4.5 + 0.7
                ws.row_dimensions[5123].height = 4.5 + 0.7
                ws.row_dimensions[5132].height = 4.5 + 0.7
                ws.row_dimensions[5134].height = 4.5 + 0.7
                ws.row_dimensions[5136].height = 4.5 + 0.7
                ws.row_dimensions[5822].height = 4.5 + 0.7
                ws.row_dimensions[5824].height = 4.5 + 0.7
                ws.row_dimensions[5826].height = 4.5 + 0.7
                ws.row_dimensions[6393].height = 4.5 + 0.7
                ws.row_dimensions[6395].height = 4.5 + 0.7
                ws.row_dimensions[6397].height = 4.5 + 0.7
                ws.row_dimensions[6460].height = 4.5 + 0.7
                ws.row_dimensions[6462].height = 4.5 + 0.7
                ws.row_dimensions[6471].height = 4.5 + 0.7
                ws.row_dimensions[6546].height = 4.5 + 0.7
                ws.row_dimensions[6547].height = 4.5 + 0.7
                ws.row_dimensions[6549].height = 4.5 + 0.7
                ws.row_dimensions[7707].height = 4.5 + 0.7
                ws.row_dimensions[7738].height = 4.5 + 0.7
                ws.row_dimensions[7740].height = 4.5 + 0.7
                ws.row_dimensions[7774].height = 4.5 + 0.7
                ws.row_dimensions[8050].height = 4.5 + 0.7
                ws.row_dimensions[8438].height = 4.5 + 0.7
                ws.row_dimensions[8560].height = 4.5 + 0.7
                ws.row_dimensions[8569].height = 4.5 + 0.7
                ws.row_dimensions[8571].height = 4.5 + 0.7
                ws.row_dimensions[8573].height = 4.5 + 0.7
                ws.row_dimensions[9259].height = 4.5 + 0.7
                ws.row_dimensions[9261].height = 4.5 + 0.7
                ws.row_dimensions[9263].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='L', end='X', hidden=False) 
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                ws.column_dimensions.group(start='BC', end='BN', hidden=True)
                ws.column_dimensions.group(start='BQ', end='CB', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'CF','CG','CH','CI','CJ','CK','CL']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:CM9937'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK9937'
                else:
                    tamanho_tb = 'A11:CM9937'
                ws.auto_filter.ref = tamanho_tb

            elif 'S_' in i:            

                # Converter colunas numéricas para porcentagem
                for a in s_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')        


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in s_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito 

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in s_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 55
                while column < 68:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 69
                while column < 82:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1   

                column = 85
                while column < 89:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1   

                column = 89
                while column < 92:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1   

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 15 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7
                ws.column_dimensions['BP'].width = 5 + 0.7
                ws.column_dimensions['CD'].width = 3 + 0.7
                ws.column_dimensions['CE'].width = 3 + 0.7


                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[660].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                ws.column_dimensions.group(start='BC', end='BN', hidden=True)
                ws.column_dimensions.group(start='BQ', end='CB', hidden=True)
                for col in ['H', 'I', 'J', 'L']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:CM972'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK972'
                else:
                    tamanho_tb = 'A11:CM972'
                ws.auto_filter.ref = tamanho_tb

            elif 'Casc_' in i:

                # Converter colunas numéricas para Contábil
                for a in  c_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                # Centralizar coluna de Cod
                for k in range(12, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('cascmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = 500, min_col = 1, max_col = 81):
                    cell_value = row[5].value
                    if cell_value in c_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito   
                    elif cell_value in c_cod_cinza_medio:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_cinza_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_escuro
                            cell.font = fonte_preta_negrito 
                    elif cell_value in c_cod_vermelho:
                        for cell in row:
                            cell.fill = backgroud_vermelho
                            cell.font = fonte_branca_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''


                # Configurar estilo header
                for n in c_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column H onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 55
                while column < 68:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1      

                column = 69
                while column < 82:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 0.7
                ws.column_dimensions['B'].width = 0.7
                ws.column_dimensions['C'].width = 0.7
                ws.column_dimensions['D'].width = 0.7
                ws.column_dimensions['E'].width = 0.7
                ws.column_dimensions['F'].width = 16 + 0.7
                ws.column_dimensions['G'].width = 50 + 0.7
                ws.column_dimensions['H'].width = 1.86 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 1.86 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 3 + 0.7
                ws.column_dimensions['AN'].width = 3 + 0.7
                ws.column_dimensions['BB'].width = 3 + 0.7
                ws.column_dimensions['BP'].width = 3 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[39].height = 4.5 + 0.7
                ws.row_dimensions[66].height = 4.5 + 0.7
                ws.row_dimensions[69].height = 4.5 + 0.7
                ws.row_dimensions[92].height = 4.5 + 0.7
                ws.row_dimensions[95].height = 4.5 + 0.7
                ws.row_dimensions[111].height = 4.5 + 0.7
                ws.row_dimensions[113].height = 4.5 + 0.7
                ws.row_dimensions[118].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                ws.column_dimensions.group(start='BC', end='BN', hidden=True)
                ws.column_dimensions.group(start='BQ', end='CB', hidden=True)
                for col in ['A', 'B', 'C', 'D']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:CC135'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BA135'
                else:
                    tamanho_tb = 'A11:CC135'
                ws.auto_filter.ref = tamanho_tb

            elif 'PNL_' in i: 
                pass
            
            elif 'anl' in i:
                ws.auto_filter.ref = ws.dimensions

            else:
                #Tab PModelo           

                # Converter colunas numéricas para Contábil
                for a in  p_cols_num:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '_-* #,##0_-;(#,##0);_-* "-"??_-;_-@_-'

                for a in  p_cols_perc:
                    for b in range(12, len(ws['F'])+1):
                        ws[f'{a}{b}'].number_format = '0.00%'

                # Centralizar coluna de Cod
                for k in range(3, len(ws['F'])+1):
                    ws[f'F{k}'].alignment = Alignment(horizontal='center', vertical='center')

                # set tab color
                ws.sheet_properties.tabColor = cor_tabs.get('pmodelo')


                # Configurar estilo linhas específicas
                for row in ws.iter_rows(min_row = 12, max_row = ws.max_row, min_col = 1, max_col = 91):
                    cell_value = row[5].value
                    if cell_value in p_cod_cinza_medio_escuro:
                        for cell in row:
                            cell.fill = backgroud_cinza_medio_escuro
                            cell.font = fonte_preta_negrito

                # Replace valor -999 para espaço vazio na coluna de Cod
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == null_value:
                            cell.value = ''

                # Configurar estilo header
                for n in p_head:
                    ws[n].alignment = Alignment(horizontal='center', vertical='center')
                    ws[n].fill = backgroud_vermelho
                    ws[n].font = fonte_branca_negrito

                ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['H11'].fill = backgroud_cinza_escuro
                ws['H11'].font = fonte_branca_negrito

                #Congelar célula
                freeze_cell = ws['L12']
                ws.freeze_panes = freeze_cell
                
                # Zoom na sheet
                ws.sheet_view.zoomScale = 70


                # Start changing width from column M onwards
                column = 13
                while column < 26:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 27
                while column < 40:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 41
                while column < 54:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 55
                while column < 68:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 69
                while column < 82:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 20.43 + 0.7
                    column += 1

                column = 82
                while column < 89:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 3 + 0.7
                    column += 1

                column = 89
                while column < 92:
                    m = get_column_letter(column)
                    ws.column_dimensions[m].width = 7 + 0.7
                    column += 1

                # set the width of the column
                ws.column_dimensions['A'].width = 27.43 + 0.7
                ws.column_dimensions['B'].width = 20.71 + 0.7
                ws.column_dimensions['C'].width = 14.29 + 0.7
                ws.column_dimensions['D'].width = 17.86 + 0.7
                ws.column_dimensions['E'].width = 16.43 + 0.7
                ws.column_dimensions['F'].width = 14.14 + 0.7
                ws.column_dimensions['G'].width = 45.57 + 0.7
                ws.column_dimensions['H'].width = 14.71 + 0.7
                ws.column_dimensions['I'].width = 1.86 + 0.7
                ws.column_dimensions['J'].width = 14.71 + 0.7
                ws.column_dimensions['K'].width = 1.86 + 0.7
                ws.column_dimensions['L'].width = 1.86 + 0.7
                ws.column_dimensions['Z'].width = 5 + 0.7
                ws.column_dimensions['AN'].width = 5 + 0.7
                ws.column_dimensions['BB'].width = 5 + 0.7
                ws.column_dimensions['BP'].width = 5 + 0.7

                # set the height of the row
                ws.row_dimensions[1].height = 4.5 + 0.7
                ws.row_dimensions[2].height = 4.5 + 0.7
                ws.row_dimensions[3].height = 4.5 + 0.7
                ws.row_dimensions[4].height = 4.5 + 0.7
                ws.row_dimensions[5].height = 4.5 + 0.7
                ws.row_dimensions[6].height = 4.5 + 0.7
                ws.row_dimensions[7].height = 4.5 + 0.7
                ws.row_dimensions[8].height = 4.5 + 0.7
                ws.row_dimensions[9].height = 4.5 + 0.7
                ws.row_dimensions[395].height = 4.5 + 0.7
                ws.row_dimensions[629].height = 4.5 + 0.7
                ws.row_dimensions[662].height = 4.5 + 0.7
                ws.row_dimensions[696].height = 4.5 + 0.7
                ws.row_dimensions[1832].height = 4.5 + 0.7
                ws.row_dimensions[2066].height = 4.5 + 0.7
                ws.row_dimensions[2099].height = 4.5 + 0.7
                ws.row_dimensions[2133].height = 4.5 + 0.7
                ws.row_dimensions[3091].height = 4.5 + 0.7
                ws.row_dimensions[3325].height = 4.5 + 0.7
                ws.row_dimensions[3358].height = 4.5 + 0.7

                # Ocultar e Agrupar colunas
                ws.column_dimensions.group(start='A', end='E', hidden=False)
                ws.column_dimensions.group(start='M', end='X', hidden=False)
                ws.column_dimensions.group(start='AA', end='AL', hidden=False)
                ws.column_dimensions.group(start='AO', end='AZ', hidden=True)
                ws.column_dimensions.group(start='BC', end='BN', hidden=True)
                ws.column_dimensions.group(start='BQ', end='CB', hidden=True)
                for col in ['H', 'I', 'J', 'L', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM']:
                    ws.column_dimensions[col].hidden= True

                #Inserir filtros
                if num_anos == 2:
                    tamanho_tb = 'A11:CM3667'
                elif num_anos == 3:
                    tamanho_tb = 'A11:BK3667'
                else:
                    tamanho_tb = 'A11:CM3667'
                ws.auto_filter.ref = tamanho_tb

                wb.save(path_excel)
